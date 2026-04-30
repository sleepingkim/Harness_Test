# -*- coding: utf-8 -*-
"""
얼굴/음성 바이오마커 논문 정리 Excel 생성 스크립트
출력: face_voice_biomarker_papers.xlsx (4개 시트)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 스타일 정의
# ============================================================
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
FACE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
VOICE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DATASET_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BODY_FONT = Font(name="맑은 고딕", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, headers, body_fill, n_rows):
    """헤더 행 + 본문 행 스타일링."""
    # 헤더
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[1].height = 36
    # 본문
    for r in range(2, n_rows + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = body_fill
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
        ws.row_dimensions[r].height = 60


def autosize_columns(ws, headers, max_width=50, min_width=10):
    """대략 자동 너비(한글 가중치 포함)."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # 헤더 길이 + 본문 최대 길이 추정
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
            v = row[0]
            if v is None:
                continue
            # 한글은 2배 가중
            est = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
            if est > max_len:
                max_len = est
        width = min(max(max_len * 0.65 + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ============================================================
# Sheet 1: 얼굴 바이오마커 (50편)
# ============================================================
FACE_HEADERS = [
    "번호", "제목", "저자", "연도", "게재지/학회", "바이오마커 유형", "구체적 바이오마커",
    "탐지 질환", "수집 방법/장비", "피험자 수", "특징 추출 방법", "모델 아키텍처",
    "주요 성능 (AUC/Acc)", "민감도", "특이도", "코드 공개", "데이터 공개",
    "증거 수준", "주요 한계점", "얼굴 영역", "이미지/영상", "검증 상태",
]

FACE_PAPERS = [
    [1, "Automatic Acne Object Detection and Acne Severity Grading Using Smartphone Images and AI (AcneDet)",
     "Huynh QT, Nguyen PH, Le HX et al.", 2022, "Diagnostics (MDPI) 12(8):1879",
     "피부/여드름", "4종 여드름 병변 객체(blackheads/whiteheads, papules/pustules, nodules/cysts, scars) + IGA 5단계",
     "여드름 (PCOS 안드로겐성 표현형 연관)", "iOS·Android 스마트폰, 다양한 조명",
     "1,572장(피부과 4명 라벨)", "Bounding box 라벨링, ImageNet 사전학습",
     "Faster R-CNN(객체 검출) + LightGBM(중증도)", "객체 mAP=0.54, 중증도 평균 AUC=0.85",
     "-", "-", "아니오", "아니오", "Moderate", "단일 데이터셋, 외부 검증 부재, 다민족 다양성 부족",
     "전체 얼굴", "이미지", "✅ 검증됨"],

    [2, "Development and Accuracy of an AI Algorithm for Acne Grading from Smartphone Photographs",
     "Seité S, Khammari A, Benzaquen M et al.", 2019, "Clin Cosmet Investig Dermatol",
     "피부/여드름", "GEA 척도, 병변 유형(comedones, papules, pustules)",
     "여드름", "스마트폰 셀피, 실생활 환경", "5,972장(피부과 2명 라벨)",
     "얼굴 검출, 정규화, 데이터 증강", "CNN (구조 비공개)",
     "전문의 일치 r=0.74", "-", "-", "아니오", "아니오", "Moderate",
     "스마트폰 이미지 품질 편차", "전체 얼굴", "이미지",
     "⚠️ 저널명 불일치 가능 (Exp Dermatol DOI 10.1111/exd.14022 vs CCID)"],

    [3, "A Deep Learning-Based Facial Acne Classification System",
     "Lim ZV, Akram F, Ngo CP et al.", 2022, "Clin Cosmet Investig Dermatol 15:851-861",
     "피부/여드름", "KAGS(Korean Acne Grading System) 4단계 중증도", "여드름",
     "임상 카메라 정면·측면", "1,213장(시리즈)",
     "얼굴 검출(Dlib), ROI 크롭, ImageNet 정규화, 회전·반전 증강",
     "ResNet-152, ImageNet 사전학습, Adam lr=1e-4",
     "Top-1 Acc≈67%, Top-2 Acc≈92%", "-", "-", "아니오", "아니오", "Moderate",
     "단일 인종, 데이터셋 소규모", "전체 얼굴", "이미지", "✅ 검증됨"],

    [4, "Evaluation of an Acne Lesion Detection and Severity Grading Model for Chinese Population in Online and Offline Healthcare Scenarios",
     "Peking Univ. AcneDGNet 팀", 2024, "Scientific Reports (Nature)",
     "피부/여드름", "여드름 병변 검출 + 중증도", "여드름(중국인)",
     "스마트폰 + 임상 카메라 (온/오프라인)", "수천 장",
     "CNN 객체 검출 + 분류 파이프라인", "AcneDGNet (CNN 기반)",
     "임상 시나리오에서 dermatologist 수준 일치", "-", "-", "아니오", "아니오",
     "Moderate", "단일 인종(중국), 외부 검증 제한", "전체 얼굴", "이미지", "✅ 검증됨"],

    [5, "Automated Grading of Acne Vulgaris by Deep Learning with CNNs",
     "Wu X, Wen C, Yang J et al.", 2019, "Skin Research and Technology",
     "피부/여드름", "Pillsbury 4단계 여드름 중증도", "여드름",
     "표준화된 임상 사진", "5,871장(중국 환자)", "ImageNet 사전학습",
     "ResNet-50", "Top-1 Acc~81%, Cohen's κ 전문의 수준", "-", "-",
     "-", "-", "Moderate", "외부 검증 필요", "전체 얼굴", "이미지", "✅ 검증됨 (PMID 31565821)"],

    [6, "DED: Diagnostic Evidence Distillation for Acne Severity Grading on Face Images",
     "익명(Pattern Recognition 게재)", 2024, "Pattern Recognition (Elsevier)",
     "피부/여드름", "병변 카운트 기반 IGA 등급", "여드름",
     "임상/셀피 사진", "다수", "Knowledge distillation",
     "Teacher-student CNN", "기존 CNN baseline 대비 정확도 향상", "-", "-",
     "일부", "-", "Moderate", "구체 수치 비공개", "전체 얼굴", "이미지", "❓ 미확인"],

    [7, "BiliScreen: Smartphone-Based Scleral Jaundice Monitoring for Liver and Pancreatic Disorders",
     "Mariakakis A, Banks MA, Phillipi L et al.", 2017, "Proc ACM IMWUT 1(2)",
     "눈/공막", "공막(sclera) 색상(노란빛) → 빌리루빈 추정",
     "황달, 간/췌장 질환", "스마트폰 + 색보정 부속품(블루박스, 종이 글래스)",
     "70명 임상 피험자", "얼굴/공막 검출, 색공간 변환, 색보정",
     "색공간 회귀(linear/nonlinear) + ML", "임상 빌리루빈 ≥3 mg/dL 검출 sens 89.7%",
     "89.7%", "96.8%", "-", "아니오(시스템 설명서 공개)", "Moderate-High",
     "색보정 부속품 필요", "공막", "이미지", "✅ 검증됨 (DOI 10.1145/3090085)"],

    [8, "Deep-Learning-Based Smartphone Application for Self-Diagnosis of Scleral Jaundice in Hepatobiliary/Pancreatic Diseases",
     "Kim G, Lee S, Park J et al.", 2021, "J Pers Med 11(9):928",
     "눈/공막", "스마트폰 공막 사진 → 총 빌리루빈 회귀",
     "황달", "스마트폰, 표준 조명/색보정", "입원 환자 사진+혈청 빌리루빈",
     "MTCNN 얼굴/눈 검출, U-Net 공막 분할, A4 화이트밸런스",
     "ResNet 기반 회귀", "혈청값과 강한 상관(r 보고)", "-", "-",
     "아니오", "아니오", "Moderate", "병원 환경 검증 한정", "공막", "이미지", "✅ 검증됨"],

    [9, "Non-Invasive Jaundice Screening Using AI: ML Analysis of Sclera and Urine Images",
     "(2025 J Clin Med 팀)", 2025, "J Clin Med (MDPI)",
     "눈/공막+소변", "공막 황달 + 소변 색상 결합", "황달",
     "스마트폰", "임상 코호트", "DT, RF, XGBoost, DeepSets, ResNet 비교",
     "DeepSets / ResNet", "황달 검출 Acc=87.1%, AUC=0.869, 빌리루빈 회귀 R²=0.782",
     "88.1% (recall)", "-", "아니오", "아니오", "Moderate",
     "단일 코호트", "공막+소변", "이미지", "✅ 검증됨 (PubMed 40364159)"],

    [10, "AI-Based Non-Invasive Bilirubin Prediction for Neonatal Jaundice Using 1D CNN",
     "(2025 Sci Rep 팀)", 2025, "Scientific Reports (Nature)",
     "피부색", "신생아 얼굴(피부) 색상 → 빌리루빈", "신생아 황달",
     "표준 조명 임상 카메라", "신생아 코호트", "Color channel 1D 신호",
     "1D CNN", "임상 빌리루빈 회귀 정확도 보고", "-", "-",
     "-", "-", "Moderate", "신생아 한정", "신생아 얼굴 피부", "이미지", "✅ 검증됨"],

    [11, "Neonatal Jaundice Detection Using a Vision Transformer-Based Deep Learning Model",
     "(2026 Sci Rep 팀)", 2026, "Scientific Reports (Nature)",
     "피부색", "신생아 얼굴 사진 → 황달 이진 분류", "신생아 황달",
     "임상 카메라", "신생아 코호트", "ViT", "Vision Transformer",
     "정확도/민감도/특이도 보고(원문 참조)", "-", "-", "-", "-", "Moderate",
     "신생아 한정", "신생아 전체 얼굴", "이미지", "✅ 검증됨 (PMC13000152, 2026 출판)"],

    [12, "FaceAge: A Deep Learning System to Estimate Biological Age from Face Photographs to Improve Prognostication",
     "Bontempi D, Schoenfeld JD, Bitterman DS et al.", 2025, "Lancet Digit Health",
     "노화/예후", "얼굴 사진 기반 생물학적 나이(FaceAge) → 예후",
     "암 예후, 생물학적 노화", "디지털 사진(공개+임상)",
     "학습 58,851명(IMDB-WIKI 56,304+UTKFace 2,547), 임상 검증 6,196 암 환자(미국·네덜란드)",
     "MTCNN 얼굴 검출+정렬, 256×256, ImageNet 사전학습 회귀(L1/L2)",
     "VGG/Inception 두 단계 CNN(localization+age regression)",
     "임상의 단독 AUC 0.74 → +FaceAge AUC 0.80; FaceAge 1년 증가당 사망 HR 유의",
     "-", "-", "코드 공개(GitHub AIM-Harvard/FaceAge)", "데이터 일부",
     "High", "인종/조명 편향, 화장·필터 영향", "전체 얼굴", "이미지", "✅ 검증됨 (PubMed 40345937)"],

    [13, "Decoding Biological Age from Face Photographs Using Deep Learning (FaceAge prelim)",
     "(2023 eLife 팀)", 2023, "eLife / PMC10516042",
     "노화/예후", "얼굴 → 생물학적 나이", "노화/예후",
     "디지털 사진", "다수", "VGG-16 fine-tuned",
     "CNN regression", "MAE ~ 4-5세", "-", "-", "모델 일부 공개", "-",
     "High", "데이터 편향", "전체 얼굴", "이미지", "✅ 검증됨 (PubMed 37745558)"],

    [14, "A Deep Learning-Based Detection of Wrinkles on Skin",
     "(Springer LNNS 팀)", 2022, "Springer LNNS",
     "피부 노화/주름", "주름 픽셀 분할/계수 → 노화·피부 건강 등급",
     "노화/피부 건강", "임상/셀피", "다수", "픽셀 분할 마스킹",
     "U-Net 변형", "Pixel-wise IoU 보고", "-", "-", "-", "-",
     "Limited-Moderate", "지표 한정", "얼굴 피부", "이미지", "✅ 검증됨"],

    [15, "Striped WriNet: Automatic Wrinkle Segmentation Based on Striped Attention Module",
     "(Biomed Signal Process Control 팀)", 2024, "Biomed Signal Process Control",
     "피부 노화/주름", "깊은 주름+얇은 잔주름 분할", "노화/피부",
     "임상 사진", "다수", "Striped Attention",
     "U-Net + Striped Attention", "분할 IoU/F1 향상", "-", "-",
     "-", "-", "Moderate", "분할 정확도 위주", "얼굴 피부", "이미지", "✅ 검증됨"],

    [16, "Evaluating Facial Dermis Aging in Healthy Caucasian Females with LC-OCT and Deep Learning",
     "(Sci Rep 2024 팀)", 2024, "Scientific Reports",
     "피부 노화", "LC-OCT 영상 기반 진피 노화 → 시간적 나이 회귀",
     "노화", "LC-OCT 영상", "100명(백인 여성 20-70세)", "3D 입력",
     "3D ResNet-18", "MAE=4.2년, Pearson r=0.93", "-", "-", "-", "-",
     "Moderate", "LC-OCT 모달리티(일반 카메라와 차별)", "얼굴 진피", "3D 영상", "❓ 미확인"],

    [17, "Detection of Anemia Using Conjunctiva Images: A Smartphone Application Approach",
     "(Smart Health 팀)", 2023, "Smart Health (Elsevier)",
     "눈/결막", "안검 결막 RGB·HSV 색상 → 헤모글로빈 추정",
     "빈혈", "스마트폰 카메라, 눈꺼풀 외번", "임상 코호트",
     "결막 ROI 분할, 색보정", "CNN + 분류기 스택",
     "Acc=92.5%", "90%", "95%", "앱 시연 공개", "-",
     "Moderate-High", "조명 변동 민감", "안검 결막", "이미지", "❓ 미확인"],

    [18, "Deep Learning Model-Based Detection of Anemia from Conjunctiva Images",
     "(HIR 2025 팀)", 2025, "Healthc Inform Res 31(1):57",
     "눈/결막", "결막 이미지 → 빈혈(이진)", "빈혈",
     "스마트폰 카메라", "764장 → DCGAN 증강 4,315장",
     "Stacking ensemble", "VGG-16 + ResNet-50 + InceptionV3 (ensemble)",
     "AUC=0.97", "-", "-", "아니오", "아니오", "Moderate",
     "GAN 증강 의존, 외부 검증 부재", "안검 결막", "이미지", "✅ 검증됨 (PubMed 39973037)"],

    [19, "Non-Invasive Anemia Detection from Conjunctiva and Sclera Images Using ViT with Attention Map Explainability",
     "(Sci Rep 2025 팀)", 2025, "Scientific Reports",
     "눈/결막+공막", "결막+공막 ROI 색상", "빈혈",
     "스마트폰 카메라", "임상 코호트", "Transfer learning",
     "Vision Transformer (ViT)", "전체 정확도 98.47%", "-", "-",
     "Attention map 설명 제공", "-", "Moderate-High",
     "단일 코호트", "결막+공막", "이미지", "✅ 검증됨"],

    [20, "AnemiaVision: Non-Invasive Anemia Detection via Smartphone Imagery Using EfficientNet-B3",
     "(IIIT Surat 팀)", 2026, "arXiv:2604.22964",
     "눈/결막", "결막 이미지", "빈혈", "스마트폰",
     "코호트 미공개", "TrivialAugmentWide + Mixup + RandomErasing",
     "EfficientNet-B3", "Validation Acc=94-97%", "-", "-",
     "코드 공개", "-", "Moderate", "검증 코호트 제한",
     "안검 결막", "이미지", "✅ 검증됨 (arXiv 실재)"],

    [21, "Prediction of Anemia in Real-Time Using a Smartphone Camera Processing Conjunctival Images",
     "(PLOS One 2024 팀)", 2024, "PLOS One",
     "눈/결막", "결막 이미지 헤모글로빈 회귀", "빈혈",
     "스마트폰 카메라", "임상 코호트", "결막 ROI 처리",
     "CNN + ML", "헤모글로빈 추정 r 보고", "-", "-",
     "-", "-", "Moderate", "표본 정보 한정", "안검 결막", "이미지", "❓ 미확인"],

    [22, "Explainable Deep Learning System for Automatic Detection of Thyroid Eye Disease Using Facial Images",
     "(Am J Ophthalmol 2025 팀)", 2025, "Am J Ophthalmol",
     "눈/안와", "안와 주변 랜드마크 + TED 활성 임상 징후",
     "갑상선안병증(TED)", "임상 정면 얼굴 사진", "다기관 코호트",
     "Periocular landmark localization network",
     "두 단계 — landmark net + binary CNN",
     "AUC=0.997", "99.7%", "94.5%", "아니오", "아니오",
     "High", "인종 편향, 화장 영향", "안와 주변", "이미지", "✅ 검증됨 (PubMed 40441501)"],

    [23, "Deep Learning-Driven Exophthalmometry Through Facial Photographs in Thyroid Eye Disease",
     "(Ophthalmol Sci 2025 팀)", 2025, "Ophthalmology Science",
     "눈/안와", "Hertel 안구 돌출 mm 회귀", "갑상선안병증(TED)",
     "얼굴 사진+Hertel 측정 페어", "코호트", "얼굴 정렬, 안와 ROI",
     "Dual-stream ResNet-18", "MAE=1.27 mm, Pearson r=0.82",
     "-", "-", "아니오", "-", "High", "단일 코호트",
     "안와", "이미지", "✅ 검증됨 (PMC12150081)"],

    [24, "Machine Learning-Assisted System Using Digital Facial Images to Predict Clinical Activity Score in Thyroid-Associated Orbitopathy",
     "Moon JH, Shin K, Lee GM et al.", 2022, "Sci Rep 12:22085",
     "눈/안와", "TAO CAS 항목별(안검부종, 결막부종/충혈, 안검 발적, 카룬쿨 부종)",
     "갑상선안병증(TAO)", "다기관 임상 정면 얼굴", "다기관",
     "ImageNet 사전학습", "ResNet-18 ensemble",
     "Active TAO 진단 sens 0.881, spec 0.869; 각 항목 AUC 0.884-0.977",
     "88.1%", "86.9%", "아니오", "아니오", "High",
     "다기관이지만 인종 편향", "안와", "이미지", "✅ 검증됨 (PubMed 36543834)"],

    [25, "Automated Computer Vision Assessment of Hypomimia in Parkinson Disease: Proof-of-Principle Pilot Study",
     "Abrami A, Gunzler S, Kilbane C et al.", 2021, "JMIR 23(2):e21037",
     "표정/얼굴 근육", "얼굴 모방·표정 비디오 hypomimia 신호",
     "파킨슨병", "셀피 비디오(YouTube 등)",
     "PD 자가식별 107명 + 대조군 1,595개 비디오",
     "OpenFace 얼굴 검출+프레임 추출", "CNN(image classification)",
     "테스트셋 AUC=0.71 (vs 신경과 의사 0.75)", "-", "-",
     "아니오", "아니오", "Moderate", "자가식별 PD, 비디오 품질 편차",
     "전체 얼굴", "비디오", "✅ 검증됨 (PubMed 33616535)"],

    [26, "Detection of Hypomimia in Patients with PD via Smile Videos",
     "(Front Neurol 2021 팀)", 2021, "Frontiers in Neurology / PMC8422154",
     "표정/얼굴 근육", "미소 동작 동안 얼굴 기하·텍스처 변화",
     "파킨슨병", "PD + 대조군 미소 비디오", "임상 코호트",
     "핸드크래프티드 기하·텍스처 특징",
     "SVM, RF, Bayesian, DT", "Acc=81-90%", "-", "-",
     "아니오", "아니오", "Moderate", "수치 범위 보고",
     "입/얼굴 하부", "비디오", "❓ 미확인"],

    [27, "Explaining Facial Action Units' Correlation with Hypomimia and Clinical Scores in Parkinson's Disease",
     "(npj Park Dis 2025 팀)", 2025, "npj Parkinson's Disease",
     "표정/얼굴 근육", "FACS AU(AU1, AU4, AU6, AU12) 강도 → MDS-UPDRS-III 상관",
     "파킨슨병", "비디오", "환자+대조군",
     "OpenFace AU 추출 + 통계", "ML",
     "임상 점수 예측 r/F1 보고", "-", "-", "일부", "-",
     "Moderate", "코호트 정보 일부 비공개", "전체 얼굴 AU", "비디오", "✅ 검증됨"],

    [28, "A Study on the Possible Diagnosis of Parkinson's Disease on the Basis of Facial Image Analysis",
     "(Electronics 2021 팀)", 2021, "Electronics 10(22):2832",
     "표정", "정면 얼굴 사진 표정 분류", "파킨슨병",
     "정면 사진", "환자+대조군", "ImageNet 사전학습",
     "VGG/Inception 변형", "분류 정확도 보고", "-", "-",
     "-", "-", "Limited-Moderate", "정지 이미지 한정",
     "전체 얼굴", "이미지", "❓ 미확인"],

    [29, "Explainable Depression Detection Based on Facial Expression Using LSTM on Attentional Intermediate Feature Fusion with Label Smoothing",
     "(Sensors 2023 팀)", 2023, "Sensors 23(23):9402",
     "표정/시간동역학", "AU intensity, 시간적 표정 동역학",
     "우울증", "AVEC 인터뷰 비디오류", "AVEC 데이터",
     "OpenFace AU + LSTM + attention fusion", "LSTM + Attention",
     "Acc=91.67%, F1=88.89%", "87.03% (recall)", "-",
     "아니오", "-", "Moderate", "AVEC 한정",
     "전체 얼굴 AU", "비디오", "✅ 검증됨 (PubMed 38067773)"],

    [30, "Automatic Identification of Depression Using Facial Images with Deep CNN",
     "(PMC9281460 팀)", 2022, "Comput Intell Neurosci / PMC9281460",
     "표정/정지 이미지", "정지 얼굴 이미지 → 우울 여부", "우울증",
     "정지 이미지", "임상 코호트", "DCNN", "Deep CNN",
     "Acc~80%", "-", "-", "-", "-", "Limited-Moderate",
     "정지 이미지 한정", "전체 얼굴", "이미지", "❓ 미확인"],

    [31, "Demystifying Mental Health by Decoding Facial Action Unit Sequences",
     "(Big Data Cogn Comput 2024 팀)", 2024, "Big Data Cogn Comput 8(7):78",
     "표정/AU 시퀀스", "AU 시퀀스 시간적 패턴", "우울/불안",
     "비디오", "다수", "AU 시퀀스 추출",
     "LSTM/Transformer on AU sequences", "다중 분류 정확도",
     "-", "-", "-", "-", "Moderate", "데이터 정보 일부 미공개",
     "전체 얼굴 AU", "비디오", "❓ 미확인"],

    [32, "FacialPulse: An Efficient RNN-Based Depression Detection via Temporal Facial Landmarks",
     "(arXiv 2408.03499 / ACM MM 2024)", 2024, "arXiv:2408.03499 / ACM MM 2024",
     "표정/랜드마크", "시간적 얼굴 랜드마크 동역학", "우울증",
     "비디오", "다수", "랜드마크 시퀀스",
     "RNN + landmark sequence", "우울 분류 SOTA 갱신",
     "-", "-", "-", "-", "Moderate", "데이터 정보 일부 미공개",
     "랜드마크", "비디오", "✅ 검증됨 (DOI 10.1145/3664647.3681546)"],

    [33, "Pain Assessment from Facial Expression: Neonatal Convolutional Neural Network (N-CNN)",
     "Zamzmi G, Pai CY, Goldgof D et al.", 2019, "IEEE EMBC (실제 IJCNN 2019)",
     "표정", "신생아 얼굴 표정(이마 주름, 눈 짜기, 비순구, 입 열림)",
     "신생아 통증", "NICU 비디오/이미지", "NICU 코호트",
     "전용 N-CNN", "전이학습 미사용 N-CNN", "Acc~91%",
     "-", "-", "아니오", "아니오", "Moderate", "NICU 한정",
     "신생아 전체 얼굴", "비디오", "⚠️ 컨퍼런스명 오기 (실제 IEEE IJCNN 2019)"],

    [34, "Human vs. Machine Learning Based Detection of Facial Weakness Using Video Analysis",
     "(Front Neurol 2022 팀)", 2022, "Frontiers in Neurology / PMC9284117",
     "얼굴 비대칭", "얼굴 가중치(asymmetry), 미소·찡그림 좌우 차이",
     "뇌졸중/안면신경마비", "공개 비디오", "환자+대조군",
     "Landmark + HoG", "SVM (랜드마크) + CNN 비교",
     "Acc=89.7% (패러메딕 동등)", "-", "-", "아니오", "-",
     "Moderate", "비디오 품질 편차", "전체 얼굴", "비디오", "✅ 검증됨"],

    [35, "DeepStroke: An Efficient Stroke Screening Framework for ER with Multimodal Adversarial Deep Learning",
     "Cai Y, Zhang Y, Cai H et al.", 2022, "Med Image Anal (Elsevier)",
     "얼굴 비대칭+음성(멀티모달)", "얼굴 동영상 미세 근육 비조정성 + 음성",
     "뇌졸중", "ER 환자 비디오·오디오", "ER 코호트",
     "얼굴 검출/정렬, 시간적 윈도우, 음성 처리",
     "Multimodal adversarial DL (facial CNN + audio + fusion)",
     "Acc=79.27%, ER 의사 동등", "93.12%", "-", "아니오", "-",
     "Moderate-High", "단일 ER", "전체 얼굴+음성", "비디오+음성", "✅ 검증됨"],

    [36, "Deep Learning-Driven Diagnosis: A Multi-Task Approach for Segmenting Stroke and Bell's Palsy",
     "(Pattern Recognition 2024 팀)", 2024, "Pattern Recognition (Elsevier)",
     "얼굴 비대칭/분할", "안면마비 부위 분할 + 원인 분류",
     "뇌졸중 vs Bell's palsy", "임상 사진", "다수",
     "분할 마스킹", "Multi-task CNN", "Acc~88%",
     "-", "-", "아니오", "-", "Moderate", "수치 범위 보고",
     "얼굴 마비 부위", "이미지", "✅ 검증됨"],

    [37, "Identifying Facial Phenotypes of Genetic Disorders Using Deep Learning (DeepGestalt / Face2Gene)",
     "Gurovich Y, Hanani Y, Bar O et al.", 2019, "Nature Medicine 25:60-64",
     "얼굴 형태/랜드마크", "얼굴 사진 형태학적 특징 → 200+ 유전 증후군",
     "유전 증후군 200+", "임상 사진", "17,000+ 사진/200+ 증후군",
     "얼굴 검출/정렬", "Deep CNN(syndrome gestalt) + SVM 후처리",
     "Top-10 sens 91% (295/323); Noonan 아형 임상의 능가",
     "91%", "-", "상용(Face2Gene)", "비공개(상용)", "High",
     "인종/연령 편향(소아 우세)", "전체 얼굴", "이미지", "✅ 검증됨"],

    [38, "Efficiency of Computer-Aided Facial Phenotyping (DeepGestalt) in Individuals With and Without a Genetic Syndrome",
     "Pantel JT, Hertzberg J, Danyel M et al.", 2020, "JMIR",
     "얼굴 형태/랜드마크", "얼굴 → 증후군 vs 비증후군",
     "유전 증후군 진단", "임상 사진", "코호트",
     "DeepGestalt + SVM", "DeepGestalt + SVM",
     "분리 능력 SVM 후 큰 폭 향상", "-", "-",
     "아니오(상용)", "-", "Moderate", "상용 시스템 의존",
     "전체 얼굴", "이미지", "✅ 검증됨"],

    [39, "Deep-Learning Approach to Automatic Identification of Facial Anomalies in Endocrine Disorders",
     "(Neuroendocrinology 팀)", 2019, "Neuroendocrinology (Karger)",
     "얼굴 형태", "얼굴 기하·텍스처 → acromegaly + Cushing",
     "Acromegaly, Cushing", "임상 사진", "환자+대조군",
     "얼굴 정렬", "CNN 분류기", "Acc=92.3-100% (전문의 능가)",
     "-", "-", "아니오", "-", "Moderate-High",
     "코호트 제한", "전체 얼굴", "이미지",
     "⚠️ 연도 오기 (epub 2019, 인쇄 Neuroendocrinology 2020;110(5):328-337)"],

    [40, "Comparative Analysis of Pre-trained DL Models and DINOv2 for Cushing's Syndrome Diagnosis in Facial Analysis",
     "(arXiv 2501.12023 팀)", 2025, "arXiv:2501.12023",
     "얼굴 형태/자가지도", "얼굴 → Cushing 이진",
     "Cushing 증후군", "임상 사진", "코호트",
     "ImageNet 사전학습 + DINOv2 자가지도 비교",
     "다중 CNN + DINOv2", "AUC=0.96+", "-", "-",
     "일부", "-", "Moderate", "단일 코호트",
     "전체 얼굴", "이미지", "✅ 검증됨"],

    [41, "Real-Time Detection of Acromegaly from Facial Images with Artificial Intelligence",
     "Kong et al.", 2023, "Endocrine (실제 Eur J Endocrinol 188(1):158-166)",
     "얼굴 형태", "코·광대·입술 thickening, 주름",
     "Acromegaly + Cushing", "임상 사진",
     "Acromegaly + 대조군", "얼굴 ID 임베딩 + 분류기 (FaRL 비교)",
     "FaRL 임베딩 + 분류기", "AUC=0.965 (Cushing), 0.956 (acromegaly), Acc=0.95-0.96",
     "-", "-", "아니오", "-", "Moderate-High", "단일 코호트",
     "전체 얼굴", "이미지", "⚠️ 저널명 오기 (Eur J Endocrinol 2023, PMID 36747333)"],

    [42, "Deep Learning Algorithm for Automated Detection of Polycystic Ovary Syndrome Using Scleral Images",
     "Lv W, Song Y, Fu R et al.", 2021, "Front Endocrinol 12:789878",
     "눈/공막", "공막 혈관·색소 패턴(중의학 영감)", "PCOS",
     "임상 카메라 정면 눈 클로즈업", "중국 여성 721명 (PCOS 388 + 대조 333)",
     "얼굴/눈 검출 → 개선 U-Net + Attention 공막 분할",
     "ResNet-18 (특징 추출) + MIL 분류기",
     "AUC=0.979, Acc=0.929", "-", "-", "아니오", "아니오",
     "Moderate", "단일 인종(중국), 외부 검증 부재, 메커니즘 불명",
     "공막", "이미지", "✅ 검증됨 (얼굴 단독 PCOS 분류 사실상 유일)"],

    [43, "Feasibility of Using Deep Learning to Detect Coronary Artery Disease Based on Facial Photo",
     "Lin S, Li Z, Fu B et al.", 2020, "Eur Heart J 41(46):4400-4411",
     "얼굴 영상/심혈관", "얼굴 사진 기반 관상동맥질환 검출",
     "관상동맥질환(CAD)", "임상 정면 사진", "다기관 코호트",
     "얼굴 검출+ImageNet", "CNN 분류기",
     "임상 의사 대비 동등 또는 우수", "-", "-", "-", "-",
     "Moderate-High", "인종 편향, 단일 모달리티",
     "전체 얼굴", "이미지", "✅ 검증됨"],

    [44, "Application and Accuracy Analysis of Different Facial Regions Based on Deep Learning in the Diagnosis of Hypertension",
     "(Sci Rep 2025 팀)", 2025, "Scientific Reports",
     "얼굴 영역/심혈관", "얼굴 영역별 고혈압 진단", "고혈압",
     "임상 사진", "코호트", "얼굴 영역 분할",
     "Region-wise CNN", "영역별 정확도 비교 보고",
     "-", "-", "-", "-", "Moderate", "단일 코호트",
     "얼굴 영역(이마, 광대, 턱 등)", "이미지", "✅ 검증됨"],

    [45, "Type 1 and Type 2 Diabetes Measurement Using Human Face Skin Region",
     "(PMC10547572 팀)", 2023, "PMC10547572",
     "피부", "얼굴 피부 영역 → 당뇨", "T1DM/T2DM",
     "임상/스마트폰", "코호트", "피부 영역 ROI",
     "ML/CNN", "분류 정확도 보고", "-", "-", "-", "-",
     "Moderate", "단일 코호트", "얼굴 피부", "이미지", "✅ 검증됨"],

    [46, "Detection of Hyperglycemia and Hypoglycemia Using Deep Learning from Facial Images",
     "(BSPC 2025 팀)", 2025, "Biomed Signal Process Control",
     "피부/색조", "얼굴 → 고/저혈당 검출", "당뇨/혈당",
     "임상 사진", "코호트", "얼굴 ROI",
     "Deep learning", "분류 정확도 보고", "-", "-",
     "-", "-", "Moderate", "혈당 변동 추적 한정",
     "전체 얼굴", "이미지", "❓ 미확인"],

    [47, "A Novel Convolutional Neural Network for the Diagnosis and Classification of Rosacea: Usability Study",
     "(JMIR Med Inform 2021 팀)", 2021, "JMIR Med Inform",
     "피부/얼굴 피부질환", "주사(Rosacea) 분류", "주사",
     "임상 사진", "코호트", "ImageNet", "CNN",
     "사용성 평가 + 분류 정확도", "-", "-", "-", "-",
     "Moderate", "단일 코호트", "얼굴 피부", "이미지", "✅ 검증됨"],

    [48, "A Deep Learning Based Framework for Diagnosing Multiple Skin Diseases (EfficientNet-B4)",
     "(Front Med 2021 팀)", 2021, "Frontiers in Medicine",
     "피부", "다수 얼굴 피부질환", "다중 피부질환",
     "임상 사진", "다중 코호트", "ImageNet",
     "EfficientNet-B4", "분류 정확도 보고", "-", "-", "-", "-",
     "Moderate", "분류 일반화 평가 한정", "얼굴 피부", "이미지", "✅ 검증됨"],

    [49, "The HAM10000 Dataset: A Large Collection of Multi-Source Dermoscopic Images of Common Pigmented Skin Lesions",
     "Tschandl, P., Rosendahl, C., Kittler, H.", 2018, "Sci Data 5:180161",
     "피부/더모스코피", "7종 색소성 피부병변", "멜라노마/skin lesion",
     "임상 더모스코피", "10,015 dermoscopic", "ISIC archive",
     "데이터셋 자체", "분류 정확도 (후속 연구별)", "-", "-",
     "데이터 공개", "공개(ISIC)", "High", "더모스코피 한정",
     "피부 병변", "이미지", "✅ 검증됨"],

    [50, "Skin Cancer Classification With Deep Learning: A Systematic Review",
     "(Cancer 2022 팀)", 2022, "Cancer / PMC9327733",
     "피부/메타", "피부암 분류 메타리뷰", "피부암",
     "다수 데이터셋", "메타", "다양",
     "다양 (CNN 메타분석)", "메타분석 통합 정확도", "-", "-",
     "-", "-", "Moderate", "메타리뷰 한정",
     "피부 병변", "이미지/더모스코피", "❓ 미확인"],
]

# ============================================================
# Sheet 2: 음성 바이오마커 (43편)
# ============================================================
VOICE_HEADERS = [
    "번호", "제목", "저자", "연도", "게재지/학회", "바이오마커 유형", "구체적 바이오마커",
    "탐지 질환", "수집 장비/환경", "피험자 수", "발화 과제", "음향 특징 목록",
    "특징 추출 도구", "모델 아키텍처", "주요 성능 (AUC/Acc)", "민감도", "특이도",
    "코드 공개", "데이터 공개", "증거 수준", "주요 한계점", "언어", "한국 논문", "검증 상태",
]

VOICE_PAPERS = [
    [1, "Voice Biomarkers as Prognostic Indicators for Parkinson's Disease Using ML Techniques",
     "Naranjo, L. et al.", 2025, "Scientific Reports (Nature)",
     "전통 음향+딥", "jitter(절대·상대·RAP·PPQ5·DDP), shimmer(local·dB·APQ3·APQ5·DDA), HNR, NHR, F0 평균/SD, MFCC1-13, DFA, RPDE, PPE",
     "파킨슨병", "AKG-C420 콘덴서, 44.1 kHz/16-bit, 방음실",
     "PD 188명 + 대조 64명 (UCI Telemonitoring 확장)",
     "지속 모음 /a/ 5초", "F0, jitter, shimmer, HNR, MFCC, DFA, RPDE, PPE",
     "Praat, MATLAB Voice Analysis Toolbox",
     "SVM(RBF), XGBoost, RF, BiLSTM",
     "BiLSTM AUC=0.97, Acc=97%; SVM Acc=92%",
     "-", "-", "GitHub 일부", "UCI 공개", "High",
     "단일 발화 과제(/a/), 다국어 일반화 미검증", "영어", "아니오", "✅ 검증됨"],

    [2, "Pre-trained CNNs Identify PD from Spectrogram Images of Voice Samples",
     "Quan, C. et al.", 2025, "Scientific Reports (Nature)",
     "딥/스펙트로그램", "멜스펙트로그램(이미지), MFCC", "파킨슨병",
     "다양한 마이크(스마트폰 포함), 44.1 kHz",
     "PD 81명 + 대조 40명, 약 480 녹음",
     "지속 모음 /a/ + 읽기", "멜스펙트로그램, MFCC", "librosa",
     "ResNet50, VGG16, InceptionV3 (전이학습)",
     "ResNet50 Acc=99%, AUC=0.98", "-", "-", "부분 공개", "-",
     "High", "데이터셋별 성능 차 큼, cross-dataset 일반화 부족",
     "영어", "아니오", "✅ 검증됨"],

    [3, "Speech-Based Parkinson's Detection Using Pre-Trained Self-Supervised ASR Models",
     "Favaro, A. et al. (실제 Sedigh Malekroodi, H. et al.)", 2025, "Bioengineering (MDPI)",
     "딥 임베딩(자가지도)", "Wav2Vec 2.0(768-dim), HuBERT", "파킨슨병",
     "다중 코호트(NeuroVoz, GITA, Italian PVS, Czech PD)",
     "PD ~220명 + 대조 ~200명 (4개 코호트)",
     "지속 모음, 읽기, 자유 발화",
     "Wav2Vec 2.0, HuBERT 임베딩",
     "Wav2Vec 2.0 (Facebook AI), supervised contrastive",
     "Wav2Vec 2.0 + Linear classifier, attention pooling",
     "Acc=97.92% (intra), AUC=0.92 (cross-corpus)",
     "-", "-", "GitHub", "-", "High",
     "Cross-language 일반화 시 성능 저하",
     "영어/스페인어/체코어/이탈리아어", "아니오",
     "⚠️ 제1저자 오기 (실제 Sedigh Malekroodi, H. et al.)"],

    [4, "Machine Learning Smart System for Parkinson Disease Classification Using Voice as Biomarker",
     "Karaman, O. et al.", 2022, "Healthc Inform Res / PMC9388925",
     "전통 음향", "F0, jitter, shimmer, HNR, MFCC, NHR, DFA",
     "파킨슨병", "표준 임상 마이크, 22 kHz",
     "UCI Parkinson 31명/195 녹음",
     "지속 모음 /a/", "22개 음향 특징", "사전 추출(UCI)",
     "RF, SVM, KNN, XGBoost, NN",
     "XGBoost Acc=95.39%", "95%", "96%", "코드 일부", "UCI 공개",
     "Moderate", "작은 데이터셋(31명), 외부 검증 없음", "영어", "아니오", "✅ 검증됨"],

    [5, "파킨슨병 환자에 대한 효과적인 음성인식 시스템",
     "김지환·이종민 외", 2022, "한국음향학회지 41(6)",
     "딥 임베딩", "Globalformer 음성 임베딩, 멜스펙", "파킨슨병(한국어)",
     "임상 환경, 16 kHz", "건강자 사전학습+한국 PD 환자 fine-tune",
     "한국어 읽기", "Globalformer 임베딩, 멜스펙",
     "Globalformer (Transformer)", "Globalformer + CTC decoder",
     "CER 22.15% (ASR)", "-", "-", "비공개", "비공개",
     "Moderate", "분류가 아닌 ASR 성능, PD 분류 정확도 미보고",
     "한국어", "예", "✅ 검증됨"],

    [6, "음성 특징에 따른 파킨슨병 분류를 위한 알고리즘 성능 비교",
     "(한국멀티미디어학회 팀)", 2016, "한국멀티미디어학회논문지",
     "전통 음향", "F0, jitter, shimmer, HNR, MFCC 등 22개", "파킨슨병",
     "UCI 데이터", "환자 23명+대조 8명/195 녹음",
     "지속 모음 /a/", "22개 음향", "Praat",
     "J48, REPTree, Naive Bayes, MLP",
     "J48 Acc=88.72%, REPTree Acc=84.62%",
     "-", "-", "-", "UCI 공개", "Moderate",
     "작은 데이터셋, 한국인 음성 미포함",
     "한국어(연구자), UCI 영어 데이터 사용", "예", "✅ 검증됨"],

    [7, "Alzheimer's Dementia Recognition through Spontaneous Speech: The ADReSS Challenge",
     "Luz, S. et al.", 2020, "INTERSPEECH 2020 (ISCA)",
     "전통+딥+언어", "비언어적(침묵 비율, 발화 길이), MFCC, eGeMAPS, ComParE 6373, 언어",
     "알츠하이머", "임상 인터뷰, 16 kHz",
     "AD 78 + 대조 78 (DementiaBank Pitt 부분집합)",
     "그림 묘사 (Cookie Theft)",
     "MFCC, eGeMAPS, ComParE 6373, 침묵 비율",
     "openSMILE, librosa", "LR, SVM, LDA, RF, CNN baseline",
     "Baseline acoustic Acc=62.5%, 최상위 시스템 Acc=89.6%",
     "-", "-", "Baseline 공개", "챌린지 공개", "High",
     "표본 156명, 그림 묘사 한정", "영어", "아니오", "✅ 검증됨"],

    [8, "Acoustic and Language Based DL Approaches for Alzheimer's Dementia Detection From Spontaneous Speech",
     "Pappagari, R. et al. (실제 Mahajan, P. & Baths, V.)", 2021, "Front Aging Neurosci",
     "딥 임베딩+언어", "x-vector, MFCC, BERT 임베딩, eGeMAPS",
     "알츠하이머", "ADReSS", "AD 78 + 대조 78",
     "Cookie Theft 그림 묘사",
     "x-vector, MFCC, eGeMAPS, BERT", "openSMILE, Kaldi (x-vector)",
     "x-vector + PLDA, BERT fine-tuning, score-level fusion",
     "Acoustic-only Acc=72.9%, 멀티모달 fusion Acc=89.6%",
     "-", "-", "코드 부분 공개", "ADReSS", "High",
     "BERT 주도 향상", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Mahajan, P. & Baths, V.)"],

    [9, "Layer-wise Analysis of Wav2Vec for Early Detection of Cognitive Decline",
     "(Int J Speech Technol 2026 팀)", 2026, "Int J Speech Technol (Springer)",
     "딥 임베딩", "Wav2Vec 1.0/2.0 layer-wise, MFCC(비교)",
     "MCI/조기 인지저하", "임상 인터뷰",
     "MCI + 대조 ~100명", "자유 발화 + 읽기",
     "Wav2Vec layer embeddings", "Wav2Vec",
     "SVM, LR on Wav2Vec embeddings",
     "Wav2Vec+SVM Acc=71%, F1=74%",
     "-", "-", "코드 부분 공개", "-", "Moderate",
     "작은 데이터셋, 외부 검증 부족", "영어", "아니오", "✅ 검증됨"],

    [10, "알츠하이머형 치매 선별을 위한 딥러닝기반 한국 노인음성의 비언어학적 모델 연구",
     "서울대학교(박사학위)", 2024, "SNU Open Repository",
     "전통+비언어", "비언어학적 음향특징(침묵 비율, 발화 길이, 휴지), MFCC, eGeMAPS",
     "알츠하이머(한국어)", "한국 노인 임상, 16 kHz",
     "한국 노인 AD + 대조군",
     "한국어 자유 발화, 읽기, 그림 묘사",
     "MFCC, eGeMAPS, 비언어학적 특징", "openSMILE, librosa",
     "CNN, LSTM, BiLSTM 기반 분류기",
     "비언어학적 특징 기반 분류 (본문 참조)",
     "-", "-", "비공개", "비공개", "Moderate",
     "한국어 노인 데이터셋 표본 제약", "한국어", "예", "✅ 검증됨"],

    [11, "Multimodal Deep Learning for Dementia Classification Using Text and Audio",
     "Ortiz-Perez, D. et al. (실제 Lin, K. & Washington, P.Y.)", 2024, "Sci Rep",
     "멀티모달", "멜스펙트로그램, BERT 텍스트 임베딩", "AD",
     "DementiaBank Pitt", "AD 208 + 대조 104",
     "Cookie Theft 그림 묘사", "멜스펙, BERT",
     "librosa, HuggingFace BERT",
     "Audio CNN + Text Transformer + 멀티모달 fusion",
     "Acc=90.36%", "-", "-", "GitHub", "DementiaBank", "High",
     "영어 화자 한정", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Lin, K. & Washington, P.Y.)"],

    [12, "Clinical Assessment and Interpretation of Dysarthria in ALS Using Attention-Based Deep Learning",
     "Stegmann, G. et al. (실제 Merler, M. & Agurto, C.)", 2025, "npj Digital Medicine",
     "딥/Attention", "멜스펙트로그램, attention weights(음소·호흡음)",
     "ALS dysarthria", "임상 녹음",
     "ALS 125명, 2,102 녹음, 3 SLP 100점 평가",
     "단어/문장 읽기 ('car', 'more' 등)",
     "멜스펙트로그램", "직접 입력",
     "Attention CNN-Transformer hybrid",
     "R²=0.92, RMSE=6.78 (severity prediction)",
     "-", "-", "부분 공개", "-", "High",
     "영어 모노링구얼, ALS만", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Merler, M. & Agurto, C.)"],

    [13, "Dysarthria Detection Based on a DL Model with a Clinically-Interpretable Layer",
     "Tu, M. et al.", 2023, "JASA Express Letters (AIP)",
     "딥/해석가능", "멜스펙, 임상 해석 layer(음소별 distortion)",
     "Dysarthria (ALS·PD·HD)", "임상 녹음, 16 kHz",
     "ALS·PD·HD + 대조군", "표준 문장 읽기",
     "멜스펙트로그램", "직접 입력",
     "CNN + 음소-인식 attention layer", "Acc≈85%, 음소별 distortion map",
     "-", "-", "부분 공개", "-", "Moderate",
     "해석성 향상 vs 정확도 트레이드오프", "영어", "아니오", "✅ 검증됨"],

    [14, "Acoustic Analysis of Voice in Huntington's Disease Patients",
     "Velasco García, M.J. et al.", 2010, "Journal of Voice (실제 인쇄 2011, 25(2):208-17)",
     "전통 음향", "F0, jitter, shimmer, HNR, formant tracking, VOT",
     "헌팅턴병", "임상 환경", "HD 13 + 대조 13",
     "지속 모음, /pa-ta-ka/, 자유 발화",
     "F0, jitter, shimmer, HNR, formants, VOT", "Praat",
     "통계 + LDA", "Acc=94% (HD vs 대조)", "-", "-",
     "비공개", "비공개", "Limited", "매우 작은 표본(n=13)",
     "스페인어 가능성", "아니오",
     "⚠️ 연도 오기 (epub 2010, 인쇄 2011)"],

    [15, "Characterizing Vocal Tremor in Progressive Neurological Diseases via Automated Acoustic Analyses",
     "Rusz, J. et al.", 2020, "Clin Neurophysiol",
     "전통 음향/떨림", "vocal tremor freq(<4Hz, 4-7Hz), F0/진폭 변동",
     "HD/ET/MSA/ALS/PSP/PD/dystonia/MS/소뇌실조",
     "임상 녹음, 44.1 kHz", "다수 환자(다중 질환)",
     "지속 모음 /a/ 5초+", "vocal tremor 주파수, F0 변동",
     "자동 vocal tremor tracker (저자 개발)", "통계 분류",
     "HD 65%, ET 50%, MSA 40%, 소뇌실조 40%, ALS 40%, PSP 25%, PD 20%, dystonia 10%, MS 8%에서 비정상 떨림",
     "-", "-", "tracker 부분 공개", "-", "High",
     "분류 아닌 prevalence 분석", "체코어 가능성", "아니오", "✅ 검증됨"],

    [16, "Depression Recognition Using Voice-Based Pre-Training Model",
     "Yang, X. et al. (실제 Huang, X. et al.)", 2024, "Scientific Reports",
     "딥 임베딩", "Wav2Vec 2.0, eGeMAPS", "우울증",
     "DAIC-WOZ", "189 인터뷰",
     "DAIC-WOZ 자유 발화", "Wav2Vec 2.0, eGeMAPS",
     "Wav2Vec 2.0 fine-tune, openSMILE",
     "Wav2Vec 2.0 + classifier head",
     "F1=0.88, RMSE=4.3 (PHQ-8 회귀)",
     "-", "-", "GitHub", "DAIC-WOZ", "High",
     "DAIC-WOZ만, 영어 한정", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Huang, X. et al.)"],

    [17, "Diagnostic Accuracy of DL Using Speech Samples in Depression: Systematic Review and Meta-Analysis",
     "Bao, R. et al. (실제 Liu, M. et al.)", 2024, "JAMIA",
     "메타분석", "MFCC, F0, eGeMAPS, ComParE",
     "우울증", "메타(25개 연구)", "통합", "다양",
     "MFCC, F0, eGeMAPS, ComParE", "다양",
     "CNN, LSTM, Transformer 메타비교",
     "Pooled sens=0.81, spec=0.81, AUC=0.87",
     "0.81", "0.81", "메타분석 데이터", "-", "High",
     "데이터셋 이질성 큼, 외부 검증 부족", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Liu, M. et al.)"],

    [18, "Automatic Depression Detection Using Smartphone-Based Text-Dependent Speech Signals: Deep CNN",
     "Lam, G. et al. (실제 Kim, A.Y. et al.)", 2023, "JMIR",
     "딥/스펙트로그램", "멜스펙, MFCC", "우울증",
     "스마트폰 (iOS/Android), 16 kHz", "자체 코호트",
     "정해진 텍스트 읽기 (text-dependent)",
     "멜스펙, MFCC", "librosa", "Deep CNN",
     "Acc=78.14%", "-", "-", "비공개", "비공개", "Moderate",
     "스마트폰 마이크 변동성, 표본 제한", "영어", "아니오",
     "⚠️ 제1저자 오기 (실제 Kim, A.Y. et al.)"],

    [19, "음성과 텍스트를 이용하여 우울증 및 자살 위험을 평가하는 인공지능 기반 임상의사결정지원시스템",
     "서울대학교(박사학위)", 2022, "SNU Open Repository",
     "전통+텍스트", "F0, MFCC, jitter, shimmer + 텍스트 특징",
     "우울증·자살위험(한국어)", "임상 인터뷰 녹음",
     "정상 33 + 경도 우울 26 + 주요 우울 34 (2차년)",
     "M.I.N.I. 한국어판", "F0, MFCC, jitter, shimmer + 한국어 NLP",
     "openSMILE, 한국어 NLP", "LR + 앙상블, NLP 분류기",
     "음성 AUC=0.806, 텍스트 AUC=0.905, 자살위험 ensemble AUC=0.800",
     "-", "-", "비공개", "비공개", "High",
     "자살 위험 빈도 낮아 임상 적용 한계", "한국어", "예", "✅ 검증됨"],

    [20, "Voice of Mind: Deep Learning Model for Depression and Anxiety Assessment From Acoustic and Lexical Vocal Biomarkers",
     "(2025 J Voice 팀)", 2025, "Journal of Voice",
     "멀티모달", "음향(MFCC, eGeMAPS) + 어휘(lexical)",
     "우울증+불안", "자체 임상", "우울+불안 vs 대조",
     "자유 발화", "MFCC, eGeMAPS, lexical",
     "openSMILE + LLM", "멀티모달 딥러닝",
     "동시 평가 모델", "-", "-", "-", "-", "Moderate",
     "표본/외부 검증 정보 한정", "영어", "아니오", "✅ 검증됨"],

    [21, "Voice Analysis as an Objective State Marker in Bipolar Disorder",
     "Faurholt-Jepsen, M. et al.", 2016, "Translational Psychiatry",
     "전통+종단", "음성 quality(창끼·기식음), spectral flux, F1-F4 formants, LPC",
     "양극성장애", "환자 스마트폰 일상 통화",
     "양극성 28명, 121일 추적",
     "일상 전화 통화", "spectral flux, formants, LPC",
     "자체 추출 파이프라인", "Random Forest",
     "조증/혼재 AUC=0.89, 우울 AUC=0.78",
     "-", "-", "비공개", "비공개", "High",
     "작은 표본, 환자 의존 마이크 품질", "덴마크어 가능성", "아니오", "✅ 검증됨"],

    [22, "Acoustic Speech Markers for Schizophrenia-Spectrum Disorders: Diagnostic and Symptom-Recognition Tool",
     "de Boer, J.N. et al.", 2021, "Psychological Medicine (실제 인쇄 2023, 53(4):1302-1312)",
     "전통 음향", "F0(mean·var), F1-F2 variability, 음절률, 침묵 비율",
     "조현병-스펙트럼", "임상 녹음",
     "조현병 86 + 대조 80 (실제 142+142)",
     "그림 묘사, 자유 발화",
     "F0, formants, 음절률, 침묵 비율",
     "Praat, 자체 스크립트", "LR, SVM",
     "Acc=81-94% (양성/음성 증상별)",
     "-", "-", "비공개", "비공개", "High",
     "단일 사이트, 네덜란드어",
     "네덜란드어", "아니오", "⚠️ 연도 오기 (epub 2021, 인쇄 2023)"],

    [23, "Identifying Diagnostic Status and Negative Symptoms of Psychosis Using CNNs",
     "(PMC12237691 팀)", 2025, "Schizophrenia Research(실제 NPP—Digital Psychiatry and Neuroscience)",
     "딥/스펙트로그램", "멜스펙트로그램", "조현병-스펙트럼/정신증",
     "임상 녹음", "조현병+대조", "자유 발화 1-3분",
     "멜스펙트로그램", "직접 입력", "CNN",
     "Acc=87.8%, AUC=0.86", "-", "-", "-", "-",
     "High", "단일 코호트", "영어", "아니오",
     "⚠️ 저널명 오기 (실제 NPP—Digital Psychiatry and Neuroscience)"],

    [24, "Identification of Psychological Stress from Speech Signal Using Deep Learning",
     "(Healthcare Analytics 2024 팀)", 2024, "Healthcare Analytics (Elsevier)",
     "전통+딥", "MFCC, pitch, eGeMAPS",
     "스트레스/불안", "스트레스 유도 + 휴식", "코호트",
     "스트레스 발화 vs 휴식", "MFCC, pitch, eGeMAPS",
     "openSMILE", "LSTM, CNN-LSTM hybrid",
     "Acc≈85%", "-", "-", "-", "-", "Moderate",
     "라벨 신뢰성", "영어", "아니오", "✅ 검증됨"],

    [25, "Coswara — A Database of Breathing, Cough, and Voice Sounds for COVID-19 Diagnosis",
     "Sharma, N. et al.", 2021, "INTERSPEECH 2020 (실제 2020 출판)",
     "데이터셋+음향", "호흡음, 기침음, 모음, 숫자 세기",
     "COVID-19", "웹 크라우드(스마트폰·웹브라우저)",
     "2,746명", "9개 표준 과제(기침, 호흡, /a/i/u/, 1-20, 빠른 호흡)",
     "MFCC, 멜스펙트로그램", "데이터셋",
     "RF baseline + 후속 DL", "Audio-MAE Coswara AUC=0.82",
     "-", "-", "데이터·코드 공개", "공개", "High",
     "자가 보고 라벨, cross-dataset 약함",
     "다국어", "아니오", "⚠️ 연도 오기 (2020 출판)"],

    [26, "The COUGHVID Crowdsourcing Dataset: A Corpus for COVID-19 Cough Sound Research",
     "Orlandic, L. et al.", 2021, "Scientific Data / arXiv 2009.11644",
     "데이터셋+기침", "기침음 spectrogram, MFCC", "COVID-19",
     "웹 크라우드", "27,550 녹음(1,156 양성), 2,000+ 의사 라벨",
     "자발적 기침", "MFCC, 멜스펙, Spectral contrast", "librosa",
     "CNN, ResNet, transfer learning",
     "AUC 0.58-0.63 (intra-COUGHVID), cross-dataset 어려움",
     "-", "-", "데이터·코드 공개", "공개", "High",
     "자가 라벨, 노이즈 환경", "다국어", "아니오", "✅ 검증됨"],

    [27, "CovidCoughNet: CNN with Pitch-Shifting Data Augmentation for COVID-19 Detection from Cough/Breath/Voice",
     "Hamdi, S. et al.", 2023, "PMC10249348",
     "딥/멀티 입력", "MFCC, 멜스펙, pitch-shifting 증강",
     "COVID-19", "Coswara + COUGHVID",
     "2,030 + 27,550 녹음", "기침, 호흡, /a/",
     "MFCC, 멜스펙", "librosa", "1D-CNN with 다중 입력 채널",
     "Acc=97% (intra-dataset)", "-", "-", "코드 공개", "공개",
     "High", "Cross-dataset 일반화 미보장",
     "다국어", "아니오", "✅ 검증됨"],

    [28, "기침소리를 이용한 코로나19 감염 진단 모델 개발",
     "한밭대학교(석사학위)", 2022, "DBpia T16613002",
     "딥/스펙트로그램", "MFCC, 멜스펙, Spectral Contrast",
     "COVID-19(한국어)", "공개+자체 수집, SNR 필터링",
     "구체 수치 본문 참조", "기침",
     "MFCC, 멜스펙, Spectral contrast", "librosa", "CNN",
     "본문 참조", "-", "-", "비공개", "공개+자체",
     "Limited", "공개 정보 한정", "한국어 연구자, 다국어 데이터",
     "예", "✅ 검증됨"],

    [29, "EfficientNet 기반 기침 소리 감지 시스템",
     "김성준 외", 2022, "earticle A409384",
     "딥/이미지", "멜스펙트로그램(이미지)", "기침 검출",
     "자체 + 공개", "자체 음성/환경음", "기침 vs 비기침",
     "멜스펙", "librosa",
     "EfficientNet (이미지 분류 전이학습)",
     "자체 녹음 Acc=76.67%", "-", "-", "-", "-",
     "Limited", "COVID-19 분류 아닌 기침 검출",
     "한국어", "예", "✅ 검증됨"],

    [30, "Respiratory Sound Classification for Crackles, Wheezes, and Rhonchi in the Clinical Field Using Deep Learning",
     "Kim, Y. et al.", 2021, "Scientific Reports",
     "딥/스펙트로그램", "멜스펙트로그램(crackles·wheezes·rhonchi)",
     "폐렴/IPF/COPD/천식/폐암/결핵/기관지확장증",
     "임상 청진기, 4 kHz",
     "871명/1,222 정상 + 696 비정상",
     "청진음(passive)", "멜스펙트로그램", "직접 입력",
     "CNN (ResNet 변형)", "Acc=86.5%, AUC=0.93, F1=0.81",
     "-", "-", "부분 공개", "-", "High",
     "단일 사이트", "영어/공통", "아니오", "✅ 검증됨"],

    [31, "Pediatric Asthma Detection with Google's HeAR Model",
     "(arXiv 2504.20124 팀)", 2025, "arXiv 2504.20124",
     "딥 임베딩(자가지도)", "HeAR(Health Acoustic Representations) 임베딩(3억 오디오 사전학습)",
     "소아 천식", "공개 소아 호흡음", "공개", "기침, 호흡음",
     "HeAR 임베딩", "HeAR + linear probe", "HeAR + linear probe",
     "융합 Acc>95%", "95.6%", "95.0%", "HeAR 모델 공개", "-",
     "High", "단일 도메인", "영어", "아니오", "✅ 검증됨"],

    [32, "Obstructive Sleep Apnea Detection Based on Sleep Sounds via Deep Learning (OSAnet)",
     "Wang, R. et al.", 2022, "Nature and Science of Sleep",
     "딥/스펙트로그램", "코골이·호흡음 멜스펙트로그램",
     "수면무호흡(OSA)", "비접촉 수면실 마이크",
     "OSA + 대조 (PSG 라벨)",
     "수면 중 비접촉 음성(전체 밤)", "멜스펙, MFCC", "직접 입력",
     "OSAnet (CNN+LSTM+DNN)",
     "호흡사건 Acc=95.3%, OSAHS 중증도 3분류 Acc=81.6%",
     "-", "-", "비공개", "비공개", "High",
     "단일 사이트", "공통", "아니오", "✅ 검증됨"],

    [33, "Voice Analysis in Women with Polycystic Ovary Syndrome",
     "Aydin, K. et al.", 2024, "Egyptian Journal of Otolaryngology (Springer)",
     "전통 음향", "F0(mean), jitter, shimmer, MPT, HNR, RAP",
     "PCOS", "임상 환경, MDVP",
     "PCOS + 대조 (수십~수백 명)",
     "지속 모음 /a/", "F0, jitter, shimmer, MPT, HNR, RAP",
     "MDVP, Praat", "통계 검정 (t-test, Mann-Whitney)",
     "PCOS 군 F0 감소·MPT 감소·RAP 증가 (유의)",
     "-", "-", "-", "-", "Moderate",
     "분류 정확도 미보고 (관찰 연구)", "영어", "아니오", "✅ 검증됨"],

    [34, "Vocal Changes in Patients With Polycystic Ovary Syndrome",
     "Aydin, K. et al.", 2010, "Journal of Voice (PubMed 20537860)",
     "전통+설문", "F0, jitter, shimmer, HNR, MPT, vocal symptom 설문",
     "PCOS", "임상 녹음", "PCOS 34 + 대조 30",
     "지속 모음, 표준 문장 읽기",
     "F0, jitter, shimmer, HNR, MPT", "MDVP", "통계",
     "PCOS throat-clearing 76.5% vs 4.8% (유의), F0 lowering",
     "-", "-", "-", "-", "Moderate",
     "ML 분류 모델 부재", "영어", "아니오", "✅ 검증됨"],

    [35, "Voice Changes in Reproductive Disorders, Thyroid Disorders and Diabetes: A Review",
     "(2022 Endocrine Connections 팀)", 2022, "Endocrine Connections / PMC8942322",
     "리뷰", "F0 변화, jitter·shimmer, HNR, vocal fold 부종/근비대",
     "갑상선/PCOS/당뇨", "리뷰", "메타", "다양",
     "F0, jitter, shimmer, HNR", "리뷰", "리뷰",
     "갑상선저하증→F0 감소·hoarseness; 갑상선항진증→tremulous; PCOS→F0 감소·voice deepening; 정상화 후 3-6개월 회복",
     "-", "-", "-", "-", "High (리뷰)",
     "리뷰 한계", "영어", "아니오", "✅ 검증됨"],

    [36, "Acoustic Analysis and Prediction of Type 2 Diabetes Mellitus Using Smartphone-Recorded Voice Segments",
     "Kaufman, J.M. et al.", 2023, "Mayo Clin Proc Digital Health",
     "전통+ML", "F0, jitter, shimmer, MFCC, formant tremor, HNR",
     "T2DM", "스마트폰(자가 녹음)",
     "267명 약 18,000+ 음성", "표준 문장 읽기 6-10초",
     "F0, jitter, shimmer, MFCC, formants", "openSMILE eGeMAPS",
     "LR + ensemble (위험요인 결합)",
     "(보고서) 여성 Acc=89%, 남성 Acc=86%; (실제) 여성 75%, 남성 70%",
     "-", "-", "데이터 일부", "데이터 일부", "Moderate",
     "자가 보고 진단, 단일 모집단", "영어", "아니오",
     "⚠️ 성능 수치 과장 (실제 75%/70%)"],

    [37, "Voice Pathology Detection on the Saarbrücken Voice Database (SVD)",
     "(Cordeiro et al. 2018, Wu et al. 2018, MDPI 2021 종합)", 2021, "다수 (대표 MDPI Appl Sci 11(15):7149)",
     "전통+딥", "MFCC, GMM with HNR/NNE/GNE, 멜스펙",
     "음성장애 다수", "SVD 표준 녹음",
     "SVD 2,043 (정상 687 + 병리 1,356; 또는 851/1,002)",
     "모음 /i, a, u/(일반·고저음), 'Guten Morgen, wie geht es Ihnen?'",
     "MFCC, HNR, NNE, GNE, 멜스펙", "openSMILE, Praat, GMM-MFCC",
     "SVM, GMM, CNN, OpenL3 transfer",
     "SVM 여 Acc=99.50%, 남 Acc=99.19%; OpenL3 transfer Acc=99.44%",
     "-", "-", "-", "공개(SVD 무료)", "High",
     "클래스 불균형, intra-DB 과적합 우려", "독일어", "아니오", "✅ 검증됨"],

    [38, "A Classification Benchmark for AI Detection of Laryngeal Cancer from Patient Voice",
     "(arXiv 2412.16267 팀)", 2024, "arXiv 2412.16267",
     "전통+딥", "MFCC, jitter, shimmer, HNR, F0",
     "후두암", "임상 녹음", "후두암 + 대조",
     "지속 모음, 자유 발화",
     "MFCC, jitter, shimmer, HNR, F0", "Praat",
     "SVM, XGBoost, LGBM, ANN, 1D-CNN, 2D-CNN",
     "Best balanced Acc=83.7%, AUROC=91.8%",
     "84.0%", "83.3%", "-", "-", "High",
     "단일 코호트", "영어", "아니오", "✅ 검증됨"],

    [39, "CNN Classifies Pathological Voice Change in Laryngeal Cancer with High Accuracy",
     "Kim, H. et al.", 2020, "J Clin Med / PMC7692693",
     "딥/스펙트로그램", "멜스펙트로그램, MFCC", "후두암",
     "임상 녹음", "후두암 + 정상", "지속 모음",
     "멜스펙, MFCC", "직접 입력", "1D-CNN, 2D-CNN",
     "1D-CNN Acc=85% (laryngologists Acc=69.9%)",
     "78%", "93%", "-", "-", "High",
     "단일 코호트", "한국어 가능성", "아니오", "✅ 검증됨"],

    [40, "Classification of Laryngeal Diseases Including Laryngeal Cancer, Benign Mucosal Disease, and Vocal Cord Paralysis by AI Using Voice Analysis",
     "(2024 Sci Rep 팀)", 2024, "Scientific Reports",
     "전통+딥", "MFCC, F0, jitter, shimmer",
     "후두암/양성 점막/성대마비/정상", "임상 녹음", "다중 분류 코호트",
     "지속 모음, 표준 문장",
     "MFCC, F0, jitter, shimmer", "Praat", "CNN",
     "5-class Acc=66.9%", "0.66", "0.91", "-", "-", "Moderate",
     "5-class 정확도 한정", "영어", "아니오", "✅ 검증됨"],

    [41, "Vocal Biomarker Is Associated With Hospitalization and Mortality Among Heart Failure Patients",
     "Maor, E. et al.", 2020, "J Am Heart Assoc",
     "딥(상용)+종단", "Beyond Verbal 음성 분석 알고리즘 (음향 패턴)",
     "심부전 (사망/입원)", "전화 음성",
     "만성 심부전 10,583명, 평균 추적 24개월",
     "30초 자유 발화 (전화 인터뷰)",
     "Beyond Verbal 임베딩", "Beyond Verbal",
     "음성 risk score (사전학습)",
     "Q4 vs Q1 사망 54% vs 23% (HR 유의)",
     "-", "-", "-", "-", "High",
     "상용 알고리즘 의존", "영어/공통", "아니오", "✅ 검증됨"],

    [42, "Infant Cry Signal Diagnostic System Using Deep Learning and Fused Features",
     "(2023 Diagnostics 팀)", 2023, "Diagnostics / PMC10297367",
     "딥/융합", "멜스펙트로그램, GFCC, HR(Hilbert-Huang Resonance), 융합",
     "신생아(청각저하/질식/갑상선저하/고빌리루빈/구개열 등)",
     "임상 NICU 녹음", "다질환 NICU 코호트",
     "신생아 울음", "멜스펙, GFCC, HR", "다중 특징 융합",
     "CNN-LSTM hybrid", "Acc=97.5% (스펙트로그램+HR+GFCC)",
     "-", "-", "-", "-", "High", "NICU 한정",
     "공통", "아니오", "✅ 검증됨"],

    [43, "AcoustoSleepMask: 수면질환 환자의 수면 중 신호 빅데이터 분석",
     "KISTI 보고서", 2021, "KISTI TRKO202100015466",
     "전통+ML", "호흡음·코골이 음향 특징",
     "수면무호흡(한국어)", "스마트 슬립 마스크",
     "보고서 본문 참조", "수면 중 음향",
     "코골이·호흡음", "기계학습 도구",
     "ML 기반 AHI 추정 알고리즘", "보고서 본문 참조",
     "-", "-", "비공개", "비공개", "Limited",
     "정부 보고서 형식", "한국어", "예", "✅ 검증됨"],
]

# ============================================================
# Sheet 3: 통합 요약
# ============================================================
SUMMARY_HEADERS = [
    "질환", "얼굴 바이오마커", "음성 바이오마커", "대표 논문", "최고 성능", "증거 수준",
    "PCOS/자궁내막증 적용 가능성",
]

SUMMARY_ROWS = [
    ["PCOS",
     "안드로겐성 여드름(IGA/GEA), 다모증(Ferriman-Gallwey), 공막 패턴(Lv 2021), FaceAge 노화",
     "F0 감소, MPT 감소, jitter/RAP 증가, throat-clearing, voice deepening",
     "얼굴: Lv 2021 (Front Endocrinol); 음성: Aydin 2010/2024 J Voice/Egypt J Otolaryngol",
     "공막 단독 AUC=0.979, 음성 통계 유의 (ML 분류 모델 부재)",
     "Moderate (얼굴 단일 코호트, 음성 ML 분류 부재)",
     "직접 적용 — 얼굴 멀티 표현형 + 한국어 음성 코호트 멀티모달이 본 하네스 핵심"],

    ["자궁내막증",
     "직접 연구 부재 (가설: 만성 통증 표정 AU, 만성 염증 피부, 노화 가속)",
     "직접 연구 부재 (가설: 호르몬 주기 F0 변동, 통증 발성 패턴)",
     "직접 논문 없음",
     "-",
     "Limited (직접 연구 부재)",
     "최우선 신규 연구 공백 — AU 통증 + 음성 호르몬 패턴 멀티모달 가능"],

    ["갑상선 질환",
     "안구 돌출(Hertel 회귀), TED CAS 항목, TED 진단(2025 AUC=0.997), 안검 부종",
     "갑상선저하: F0 감소·hoarseness·vocal fold 부종; 갑상선항진: tremulous",
     "얼굴: Am J Ophthalmol 2025, Moon 2022 Sci Rep; 음성: Endocrine Connections 2022",
     "TED Acc AUC=0.997, sens 99.7%/spec 94.5%; 갑상선 정상화 후 3-6개월 음성 회복",
     "High (얼굴), Moderate (음성)",
     "PCOS와 동반 가능성 → 멀티모달 얼굴+음성 통합 평가"],

    ["빈혈",
     "안검 결막 RGB/HSV → Hb (Sci Rep 2025 ViT 98.47%, HIR 2025 AUC=0.97, EfficientNet-B3 94-97%)",
     "직접 연구 없음",
     "Sci Rep 2025 ViT, HIR 2025, AnemiaVision arXiv 2026",
     "Acc 98.47%, AUC=0.97",
     "High (얼굴 단일 모달리티)",
     "PCOS 동반 빈혈 시 추가 가치, 자궁내막증 만성 출혈 빈혈 모니터링 가능"],

    ["황달",
     "공막 색상 → 빌리루빈 (BiliScreen sens 89.7%, Kim 2021, DeepSets R²=0.782, ViT)",
     "신생아 울음 분류 (2023 Diagnostics 다질환 Acc=97.5%)",
     "Mariakakis 2017, Kim 2021, Sci Rep 2026 ViT",
     "신생아 1D-CNN 회귀, 성인 sens 89.7%/spec 96.8%",
     "High (얼굴), Moderate (음성)",
     "-"],

    ["파킨슨병",
     "Hypomimia 비디오(Abrami 2021 AUC=0.71), FACS AU(2025), 미소 비디오 ML 81-90%",
     "지속 모음 + BiLSTM (Naranjo 2025 AUC=0.97), Wav2Vec 2.0 (Favaro 2025 cross-corpus AUC=0.92), 멜스펙+ResNet50 (Quan 2025 Acc=99%)",
     "Abrami 2021 JMIR; Naranjo 2025 Sci Rep; Favaro 2025 Bioengineering",
     "음성 Acc 99% (intra), AUC=0.92 (cross); 얼굴 AUC=0.71",
     "High (양 모달리티)",
     "직접 연관 적으나 멀티모달 신경계 평가 인프라 참조"],

    ["알츠하이머/치매",
     "FaceAge 노화 간접 지표 (직접 연구 미흡)",
     "ADReSS 멀티모달 89.6%, Wav2Vec layer-wise (MCI Acc=71%), 한국 노인 비언어학적 (서울대 2024)",
     "Luz 2020 INTERSPEECH, Pappagari/Mahajan 2021, 서울대 2024",
     "멀티모달 Acc=89.6%, MCI Acc=71%",
     "High (음성), Limited (얼굴)",
     "PCOS/자궁내막증 노년기 인지 영향 추적 가능성"],

    ["우울증",
     "AU 시퀀스 LSTM (Sensors 2023 F1=88.89%), FacialPulse 랜드마크 동역학",
     "DAIC-WOZ Wav2Vec 2.0 F1=0.88, JAMIA 2024 메타 AUC=0.87, 음향+어휘 fusion",
     "Sensors 2023; Yang/Huang 2024 Sci Rep; 서울대 박사 2022 (한국어)",
     "F1=0.88-91.67%, 멀티모달이 단일 모달리티 능가",
     "High (양 모달리티, 멀티모달 우세)",
     "PCOS/자궁내막증 동반 정신건강 평가에 직접 활용"],

    ["당뇨",
     "얼굴 피부 영역 ML (PMC10547572), 고/저혈당 얼굴 (BSPC 2025)",
     "T2DM 스마트폰 음성 + 위험요인 결합 (Mayo 2023, 실제 여 75%/남 70%)",
     "Kaufman 2023, PMC10547572",
     "T2DM Acc 70-75% (음성 단독)",
     "Moderate (양 모달리티)",
     "PCOS 인슐린 저항성 → 동반 평가 가능"],
]

# ============================================================
# Sheet 4: 공개 데이터셋
# ============================================================
DATASET_HEADERS = [
    "데이터셋명", "모달리티", "대상 질환", "데이터 규모", "수집 방법", "언어", "공개 여부", "접근 방법",
]

DATASETS = [
    ["HAM10000", "얼굴/피부 이미지(더모스코피)", "7종 색소성 피부병변(멜라노마 등)",
     "10,015 dermoscopic images", "임상 더모스코피", "-", "공개", "ISIC archive (https://www.nature.com/articles/sdata2018161)"],
    ["ISIC 2017/2018/2019", "더모스코피 이미지", "멜라노마/skin lesion",
     "25,000+ images", "임상 더모스코피", "-", "공개", "ISIC challenge 사이트"],
    ["IMDB-WIKI", "얼굴 이미지", "나이/성별",
     "500K+", "인터넷 크롤링(연예인)", "-", "공개", "https://data.vision.ee.ethz.ch/cvl/rrothe/imdb-wiki/"],
    ["UTKFace", "얼굴 이미지", "나이/성별/인종",
     "23,000+", "인터넷 크롤링", "-", "공개", "UTKFace.git"],
    ["Xiangya-Derm", "얼굴 이미지", "6종 얼굴 피부병",
     "2,656", "임상 얼굴 카메라", "-", "일부 공개", "Wu et al., IEEE Access"],
    ["AVEC (Audio-Visual Emotion Challenge)", "비디오+오디오", "우울/감정",
     "다수", "인터뷰 비디오", "-", "연구 신청", "http://avec2014.cs.nott.ac.uk/"],
    ["Face2Gene 코호트", "얼굴 이미지", "200+ 유전 증후군",
     "17,000+", "임상 얼굴 사진", "-", "비공개(상용)", "https://www.face2gene.com/"],
    ["FaceAge / AIM-Harvard", "얼굴 이미지", "생물학적 나이/예후",
     "58,851(학습) + 6,196(검증)", "공개+임상", "-", "모델 공개, 데이터 일부", "https://github.com/AIM-Harvard/FaceAge"],
    ["N-CNN NICU dataset (iCOPE)", "신생아 얼굴 비디오", "신생아 통증",
     "수백명", "NICU 비디오", "-", "일부 공개", "iCOPE/NICU 류"],
    ["AcneDet 1572", "얼굴 이미지", "여드름 등급",
     "1,572", "스마트폰", "-", "비공개", "MDPI Diagnostics 2022"],
    ["TAO multicenter", "얼굴 이미지", "갑상선안병증 활성도",
     "다기관", "임상 얼굴 사진", "-", "비공개", "Sci Rep 2022"],
    ["Conjunctiva 764→4315 (DCGAN 증강)", "얼굴/안검 결막 이미지", "빈혈",
     "4,315(증강)", "스마트폰", "-", "비공개", "HIR 2025"],
    ["UCI Parkinson Voice", "음성", "PD",
     "31명, 195 녹음", "임상실 콘덴서 마이크", "영어", "공개", "UCI ML Repo"],
    ["UCI Parkinson Telemonitoring", "음성", "PD",
     "42명, 5,875 녹음", "가정 환경", "영어", "공개", "UCI ML Repo"],
    ["mPower (Sage Bionetworks)", "음성+센서", "PD",
     "9,520명+ sustained vowel", "iPhone (ResearchKit)", "영어", "공개(등록 후)", "Synapse"],
    ["Saarbrücken Voice DB (SVD)", "음성", "음성장애 다수",
     "2,043명", "임상실", "독일어", "공개(무료)", "Saarland Univ"],
    ["DementiaBank Pitt Corpus", "음성+텍스트", "AD/MCI",
     "환자 208 + 대조 104 + 미진단 85", "임상실 인터뷰", "영어", "학술 신청", "TalkBank"],
    ["ADReSS / ADReSSo Challenge", "음성", "AD",
     "156명 + 추가", "임상실", "영어", "챌린지", "Edinburgh"],
    ["DAIC-WOZ / E-DAIC", "음성+영상+텍스트", "우울증·PTSD·불안",
     "189 / 275 인터뷰", "WoZ 인터뷰", "영어", "학술 신청", "USC ICT"],
    ["Coswara", "음성(기침·호흡·모음)", "COVID-19",
     "2,746명", "웹 크라우드(스마트폰·웹브라우저)", "다국어", "공개", "IISc Bangalore"],
    ["COUGHVID", "기침음", "COVID-19",
     "27,550 녹음", "웹 크라우드", "다국어", "공개", "EPFL"],
    ["ICBHI 2017 호흡음", "청진음", "COPD/천식/폐렴 등",
     "920 녹음, 126명", "청진기", "-", "공개", "ICBHI"],
    ["AIHub 구음장애 음성인식", "음성", "구음장애(한국어)",
     "1,200+명, 5,000-5,250시간", "임상·재활 환경", "한국어", "공개", "AIHub Korea (https://aihub.or.kr/aihubdata/data/view.do?dataSetSn=608)"],
]

# ============================================================
# Workbook 생성
# ============================================================
wb = Workbook()

# 첫 시트는 default — 이름 변경
ws1 = wb.active
ws1.title = "얼굴 바이오마커"
ws1.append(FACE_HEADERS)
for row in FACE_PAPERS:
    ws1.append(row)
style_sheet(ws1, FACE_HEADERS, FACE_FILL, len(FACE_PAPERS))
autosize_columns(ws1, FACE_HEADERS, max_width=55)

# Sheet 2: 음성 바이오마커
ws2 = wb.create_sheet(title="음성 바이오마커")
ws2.append(VOICE_HEADERS)
for row in VOICE_PAPERS:
    ws2.append(row)
style_sheet(ws2, VOICE_HEADERS, VOICE_FILL, len(VOICE_PAPERS))
autosize_columns(ws2, VOICE_HEADERS, max_width=55)

# Sheet 3: 통합 요약
ws3 = wb.create_sheet(title="통합 요약")
ws3.append(SUMMARY_HEADERS)
for row in SUMMARY_ROWS:
    ws3.append(row)
style_sheet(ws3, SUMMARY_HEADERS, SUMMARY_FILL, len(SUMMARY_ROWS))
autosize_columns(ws3, SUMMARY_HEADERS, max_width=60)

# Sheet 4: 공개 데이터셋
ws4 = wb.create_sheet(title="공개 데이터셋")
ws4.append(DATASET_HEADERS)
for row in DATASETS:
    ws4.append(row)
style_sheet(ws4, DATASET_HEADERS, DATASET_FILL, len(DATASETS))
autosize_columns(ws4, DATASET_HEADERS, max_width=55)

# 저장
output = "face_voice_biomarker_papers.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"  - Sheet 1 (Face Biomarkers): {len(FACE_PAPERS)} rows")
print(f"  - Sheet 2 (Voice Biomarkers): {len(VOICE_PAPERS)} rows")
print(f"  - Sheet 3 (Integrated Summary): {len(SUMMARY_ROWS)} rows")
print(f"  - Sheet 4 (Public Datasets): {len(DATASETS)} rows")
