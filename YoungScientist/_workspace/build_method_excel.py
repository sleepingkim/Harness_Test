"""
Phase 2 기법 분석 결과를 엑셀(.xlsx) 2시트로 정리
- Sheet1: 20개 선별 기법의 5차원 평가 + 에센스 + 축구 전이
- Sheet2: 4개 조합 파이프라인 요약
- Sheet3: 4개 연구 공백 × 기법 매핑
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (ID, 기법명, 카테고리, 원논문ID, 핵심아이디어, 혁신포인트, 축구전이성, 정확도, 가림강건성, 실시간성, 단일카메라, 조합성, 총점, 축구전이도전과제, 예상성능)
methods = [
    ("M1", "Repulsion Loss", "A. Detection", "P2-8",
     "RepGT+RepBox loss로 인접 객체 혼동 감소",
     "NMS 이전 단계에서 인접 객체 경계를 학습에 주입",
     "★★★★", 4, 5, 5, 5, 5, 24,
     "보행자는 수직 가림 / 축구는 수평·비스듬 가림 → RepGT threshold 재튜닝. 공 탐지는 RepBox를 클래스별 분리.",
     "SoccerNet Det mAP +1.5~3%, missed det -10~20%"),

    ("M2", "Head-focus Joint Detector", "A. Detection", "P2-12",
     "머리+몸 joint detection + SoftNMS",
     "가려지지 않은 머리를 body 박스의 보조 단서로 활용",
     "★★★", 4, 5, 3, 5, 4, 21,
     "방송 롱샷에서 머리 해상도 작음 → head proposal 재학습. 헤더 동작에서 head-to-head repulsion 추가.",
     "홈경기 고정카메라 +3~5%, 방송 원거리 +1%"),

    ("M3", "Occlusion-Aware Spatial Attn Transformer", "A. Detection", "P2-13",
     "Spatial attention으로 가림 영역 억제",
     "Attention 자체를 가림 감지에 활용, plug-in 가능",
     "★★★", 3, 4, 3, 5, 5, 20,
     "Detector/Re-ID 양쪽 plug-in 가능. 계산 비용 증가 유의.",
     "정성적 개선, 정량 수치 제한적"),

    ("M4", "OC-SORT", "B. Motion", "P2-5",
     "Observation-centric re-update, virtual trajectory 재구성",
     "가림 구간 Kalman 오차 누적 문제 해결",
     "★★★★★", 5, 5, 5, 5, 5, 25,
     "즉시 적용 가능. DanceTrack 등 비선형 동작 검증됨 → 축구 fast motion에 적합.",
     "SoccerNet HOTA +3~5% 기대, 700+ FPS CPU"),

    ("M5", "OATrack", "B. Motion", "P2-11",
     "Occlusion rate 추정 → adaptive Kalman gain/cue weight",
     "가림 상태를 명시적 신호로 사용",
     "★★★★★", 4, 5, 4, 5, 5, 23,
     "Occlusion rate 추정기 학습 필요. 축구 기울어진 가림에 retrain.",
     "IDF1 +2~4% 기대"),

    ("M6", "BoT-SORT CMC", "B. Motion", "P2-4",
     "Camera Motion Compensation + 확장 Kalman state",
     "방송 pan/tilt/zoom에서 association 강건",
     "★★★★", 4, 3, 4, 5, 5, 21,
     "Broadcast 축구에 즉시 유효. 홈 고정카메라에는 CMC 불필요.",
     "Broadcast IDF1 +1~3%"),

    ("M7", "ByteTrack BYTE", "C. Association", "P2-3",
     "모든 detection box 활용 (low-score 부활)",
     "가려진 객체도 low-score로 추적 유지",
     "★★★★★", 5, 5, 5, 5, 5, 25,
     "즉시 적용. 낮은 threshold 설정이 축구 ball 탐지에도 유리.",
     "SoccerNet HOTA +2~4%, IDF1 +3~5%"),

    ("M8", "GCN Association", "C. Association", "P2-14",
     "GCN affinity + pose feature",
     "가림 시 appearance 대체로 graph 관계 활용",
     "★★★★", 4, 4, 3, 5, 4, 20,
     "Graph 구성 비용 크나 정확도 높음. 실시간 요구 시 경량화 필요.",
     "MOT17 HOTA 65 수준"),

    ("M9", "Occlusion-Related GCN", "C. Association", "P2-15",
     "Graph edge에 occlusion attribute",
     "가림 관계를 명시적 모델링",
     "★★★★", 4, 5, 3, 5, 4, 21,
     "M8과 유사하나 occlusion edge 특화. 축구 amodal mask와 결합하면 강력.",
     "가림 구간 IDF1 +3~5%"),

    ("M10", "Deep-EIoU", "C. Association", "P1-11",
     "Expansion IoU + deep Re-ID (motion-agnostic)",
     "빠른 비선형 동작에도 IoU 매칭 유지",
     "★★★★★", 5, 4, 4, 5, 5, 23,
     "SportsMOT/SoccerNet에서 SOTA 검증. 즉시 baseline으로 사용.",
     "SoccerNet HOTA 85.4% 보장"),

    ("M11", "DeepSORT", "D. Re-ID", "P2-2",
     "CNN appearance + cascade matching",
     "MOT 표준 baseline",
     "★★★", 3, 3, 4, 5, 5, 20,
     "기본 baseline. 더 강한 Re-ID로 대체 권장.",
     "기준선"),

    ("M12", "Deep OC-SORT", "D. Re-ID", "P2-7",
     "OC-SORT + adaptive Re-ID weight",
     "Appearance confidence 동적 통합",
     "★★★★★", 5, 5, 4, 5, 5, 24,
     "OC-SORT + adaptive Re-ID = 장기 가림 복구 우수.",
     "MOT17/20 HOTA SOTA, 축구 적용 시 유망"),

    ("M13", "PGFA", "D. Re-ID", "P4-5",
     "Pose keypoint로 가림 영역 attention 억제",
     "가림된 part 마스킹 후 visible part attention",
     "★★★★", 4, 5, 3, 5, 5, 22,
     "축구 pose estimator 품질 중요. HRNet 등 사용 시 효과적.",
     "Occluded Re-ID mAP 개선, 축구 Re-ID +3~5%"),

    ("M14", "Sports Re-ID Part+Pose+Multi-task", "D. Re-ID", "P4-2,P4-3,P4-4",
     "Part-based + team-aware + pose alignment + multi-task",
     "축구 도메인 특화 Re-ID 종합판",
     "★★★★★", 5, 4, 3, 5, 5, 22,
     "이미 축구 Re-ID SOTA. Jersey embedding과 결합 가능성 큼.",
     "mAP 86.0, R1 81.5 (SoccerNet 2022)"),

    ("M15", "UOAIS Hierarchical Amodal", "E. Amodal", "P2-16",
     "Visible/Amodal/Occlusion 3-mask 계층 예측",
     "가려진 부분까지 복원",
     "★★★", 3, 5, 2, 5, 4, 19,
     "축구 amodal annotation 부재 → pseudo-label 필요. 학습 비용 큼.",
     "신규성 높음, 실험 난이도 ★★★★"),

    ("M16", "SAMEO", "E. Amodal", "P2-17",
     "SAM foundation + Amodal-LVIS 300K",
     "Zero-shot amodal으로 pseudo-label 가능",
     "★★★", 4, 5, 2, 5, 4, 20,
     "SoccerNet 영상에 zero-shot pseudo-label 생성 후 UOAIS로 finetune 전략.",
     "Pseudo-label 품질 70~85% 기대"),

    ("M17", "Sequential Amodal Diffusion", "E. Amodal", "P2-18",
     "Diffusion iterative + cumulative mask + uncertainty",
     "가림 구간 복원과 불확실성 동시 포착",
     "★★★", 3, 5, 1, 5, 5, 19,
     "추론 속도 매우 느림(offline). Uncertainty를 Re-ID/Jersey로 전파 가능.",
     "Uncertainty propagation 신규 기여 가능"),

    ("M18", "StrongSORT + AFLink + GSI", "F. Post-processing", "P2-6",
     "Offline tracklet linking + Gaussian interpolation",
     "중기 가림 구간 tracklet 연결 + 부드러운 궤적",
     "★★★★★", 5, 5, 4, 5, 5, 24,
     "후처리만으로 큰 개선. 축구 full-match에 즉시 적용 가능.",
     "HOTA +1~3%, IDF1 +2~5%"),

    ("M19", "GTA-Link", "F/G. Post+Integrated", "P1-12",
     "Appearance + spatio-temporal global tracklet clustering",
     "Full-match 수준 global linking",
     "★★★★★", 5, 5, 4, 5, 5, 24,
     "SoccerTrack 2025 우승 검증. 확장하여 multi-modal distance 적용 가능.",
     "SoccerNet HOTA +3~7%"),

    ("M20", "SoccerNet GSR Integrated", "G. Integrated", "P4-6,P4-7,P4-3,P3-6,P3-4",
     "통합 파이프라인 (detect+track+Re-ID+jersey+team)",
     "End-to-end Game State Reconstruction",
     "★★★★★", 5, 4, 3, 5, 5, 22,
     "가장 완성도 높은 축구 파이프라인. 개별 모듈 교체로 연구 가능.",
     "GS-HOTA SOTA, +1~3% 개선 여지"),
]

# 조합 파이프라인
combinations = [
    ("조합 1. 경량-실시간형 (Real-time Broadcast)",
     "YOLOv8 + Repulsion Loss → OC-SORT + ByteTrack + CMC → Sports Re-ID + Team color → Single-Stage Uncertainty Jersey",
     "Repulsion 감소 + ByteTrack low-score 부활 + OC-SORT 가림 복구 + CMC pan/tilt + Re-ID long-term",
     "SoccerNet HOTA 80~82%, 25~30 FPS (V100)", "★★", "V100 1장, SoccerNet-Tracking + SportsMOT pretrain"),

    ("조합 2. 고성능-오프라인형 (Post-match Analysis)",
     "RT-DETR+Repulsion+Head-focus → Deep-EIoU+BoT-SORT → PGFA+Multi-task Re-ID → GTA-Link+AFLink+GSI → Keyframe Jersey + Amodal(선택)",
     "Online + offline 이중 구조, amodal이 가림 재구성",
     "HOTA 85~88%, GS-HOTA SOTA+1~3%", "★★★★", "A100×2, SoccerNet-Tracking+GSR, 자체 amodal pseudo-label"),

    ("조합 3. 신규성 강조형 (Amodal + Uniform-aware + Uncertainty)",
     "YOLOv8+Repulsion → SAMEO Amodal → OATrack(occ rate from amodal) → Occlusion-Related GCN → Uniform-aware Re-ID(PGFA+jersey region) → Uncertainty propagation → GTA-Link(uncertainty-weighted)",
     "공백 ①②③ 동시 해결, 축구 최초 amodal 통합",
     "심각한 가림 구간 IDF1 +5~10%", "★★★★★", "A100×4, amodal 일부 annotation+검수, 긴 학습 일정"),

    ("조합 4. Long-term Identity Consistency형 (공백 ④)",
     "Deep-EIoU+ByteTrack → AFLink+GSI → Tracklet galleries (jersey+team+pose+appearance) → GTA-Link multi-modal → 22 ID cap + Hungarian global",
     "Multi-modal gallery로 full-match 일관성",
     "45min clip IDF1 60→75%+", "★★★★", "A100×2, SoccerNet 45min clip + 자체 annotation"),
]

# 연구 공백 × 기법 매핑
gaps = [
    ("공백 ① 축구 amodal perception 부재",
     "축구에서 가려진 번호·몸통의 완전 형태 복원 연구 부재",
     "UOAIS (M15), SAMEO (M16), Sequential Amodal Diffusion (M17)",
     "SAMEO로 pseudo-label 생성 → UOAIS-like head로 finetune → pose prior 투입"),

    ("공백 ② Uniform-aware occlusion disambiguation",
     "같은 팀 선수 가림 시 team color만으로는 구별 불가",
     "Sports Re-ID Part-based (M14), Multi-task Re-ID (P4-3), PGFA (M13), Contrastive Team (P4-9)",
     "Part + pose + jersey region + intra-team contrastive loss 결합"),

    ("공백 ③ Tracklet-level uncertainty propagation",
     "가림 구간 confidence를 downstream(jersey, GSR)으로 전파하는 end-to-end 부재",
     "Single-Stage Uncertainty Jersey (P3-6), OATrack (M5), Amodal Diffusion (M17), Deep OC-SORT adaptive (M12)",
     "5종 uncertainty를 end-to-end graph로 전파하는 프레임워크"),

    ("공백 ④ Long-term (full-match) identity consistency",
     "45분 전체 선수 ID 일관성 부재, 대부분 30초 clip에 국한",
     "GTA-Link (M19), AFLink+GSI (M18), Multi-task Re-ID (P4-3), Jersey majority vote (P3-7), Contrastive Team (P4-9)",
     "Multi-modal gallery + global linking + 22 ID cap"),
]

# 엑셀 생성
wb = Workbook()

# ---------- Sheet 1: 20 Methods ----------
ws1 = wb.active
ws1.title = "1. Methods (20)"

headers = ["ID", "기법명", "카테고리", "원 논문", "핵심 아이디어", "혁신 포인트", "축구 전이성",
           "정확도", "가림강건성", "실시간성", "단일카메라", "조합성", "총점",
           "축구 전이 도전과제", "예상 성능"]

for col_idx, h in enumerate(headers, start=1):
    c = ws1.cell(row=1, column=col_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cat_colors = {
    "A. Detection": "DDEBF7",
    "B. Motion": "FFF2CC",
    "C. Association": "E2EFDA",
    "D. Re-ID": "FCE4D6",
    "E. Amodal": "EDEDED",
    "F. Post-processing": "D9E1F2",
    "F/G. Post+Integrated": "D9E1F2",
    "G. Integrated": "F4CCCC",
}
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, m in enumerate(methods, start=2):
    cat = m[2]
    fill = cat_colors.get(cat, "FFFFFF")
    for col_idx, val in enumerate(m, start=1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.font = Font(size=10)
        if col_idx == 13:  # 총점 컬럼
            c.font = Font(size=11, bold=True)
            if val >= 24:
                c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

col_widths_1 = {"A":7, "B":28, "C":18, "D":14, "E":35, "F":32, "G":11,
                "H":9, "I":11, "J":10, "K":11, "L":9, "M":8,
                "N":48, "O":32}
for col, w in col_widths_1.items():
    ws1.column_dimensions[col].width = w

ws1.row_dimensions[1].height = 40
for r in range(2, len(methods) + 2):
    ws1.row_dimensions[r].height = 90

ws1.freeze_panes = "C2"
ws1.auto_filter.ref = f"A1:O{len(methods)+1}"

# ---------- Sheet 2: Combinations ----------
ws2 = wb.create_sheet("2. Pipeline 조합 (4)")
hdr2 = ["조합명", "파이프라인 구성", "기대 시너지", "예상 성능", "구현 난이도", "필요 자원"]
for col_idx, h in enumerate(hdr2, start=1):
    c = ws2.cell(row=1, column=col_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

comb_colors = ["E2EFDA", "FFF2CC", "F4CCCC", "DDEBF7"]
for row_idx, combo in enumerate(combinations, start=2):
    fill = comb_colors[(row_idx - 2) % len(comb_colors)]
    for col_idx, val in enumerate(combo, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.font = Font(size=10)

col_widths_2 = {"A":36, "B":55, "C":45, "D":35, "E":14, "F":40}
for col, w in col_widths_2.items():
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 32
for r in range(2, len(combinations) + 2):
    ws2.row_dimensions[r].height = 140

# ---------- Sheet 3: Gaps ----------
ws3 = wb.create_sheet("3. 연구 공백 × 기법")
hdr3 = ["연구 공백", "상세 설명", "적용 가능 기법", "결합 전략"]
for col_idx, h in enumerate(hdr3, start=1):
    c = ws3.cell(row=1, column=col_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

gap_colors = ["FCE4D6", "E2EFDA", "FFF2CC", "DDEBF7"]
for row_idx, g in enumerate(gaps, start=2):
    fill = gap_colors[(row_idx - 2) % len(gap_colors)]
    for col_idx, val in enumerate(g, start=1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.font = Font(size=10)

col_widths_3 = {"A":36, "B":42, "C":48, "D":50}
for col, w in col_widths_3.items():
    ws3.column_dimensions[col].width = w
ws3.row_dimensions[1].height = 32
for r in range(2, len(gaps) + 2):
    ws3.row_dimensions[r].height = 110

# ---------- Sheet 4: Summary ----------
ws4 = wb.create_sheet("4. Summary")
ws4["A1"] = "Phase 2 기법 심층 분석 결과 요약"
ws4["A1"].font = Font(bold=True, size=14)
ws4.append([])
ws4.append(["총 선별 기법 수", len(methods), "(51편 중 필터링)"])
ws4.append(["총 조합 제안", len(combinations)])
ws4.append(["연구 공백 매핑", len(gaps)])
ws4.append([])
ws4.append(["총점 상위 기법 (24점+)"])
for m in sorted(methods, key=lambda x: -x[12])[:6]:
    ws4.append([m[0], m[1], m[2], m[12]])
ws4.append([])
ws4.append(["권장 연구 방향"])
ws4.append(["1순위", "방향 A: Amodal + Uniform-aware Soccer Tracker (조합 3)"])
ws4.append(["보조", "방향 B: Occlusion-Aware Tracklet Uncertainty Propagation"])
ws4.append(["대안", "방향 C: Long-term Identity Consistency via Multi-modal Gallery"])

for col in ["A", "B", "C", "D"]:
    ws4.column_dimensions[col].width = 28
ws4.column_dimensions["B"].width = 55

# 저장
out = r"C:\Users\neohc\Desktop\ClaudeCode\YoungScientist\_workspace\04_method_analysis_table.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
