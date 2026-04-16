"""
Phase 1 문헌 탐색 결과를 Excel (.xlsx)로 변환
입력: 01_literature_review.md 내 48편 논문
출력: 02_literature_table.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 48편 논문 데이터: (ID, Stage, 저자, 제목, 연도, 학회/저널, 문제정의, 가설/접근법, 모델/아키텍처, 데이터셋, 핵심성능, 실험방법, 결론/기여점, 가림관련성, URL)
papers = [
    # ---------- Stage 1: 축구/스포츠 MOT ----------
    ("P1-1", "1. 축구 MOT", "Cioppa, Giancola et al.", "SoccerNet-Tracking: Multiple Object Tracking Dataset and Benchmark in Soccer Videos", 2022, "CVPR 2022 Workshop (CVSports)",
     "축구 영상 다중 객체 추적을 위한 대규모 공개 벤치마크 부재", "200×30초 + 1×45분 하프타임 시퀀스, 8개 클래스 주석으로 공식 벤치마크 구축",
     "ByteTrack/StrongSORT 기반 baseline", "SoccerNet-Tracking (1080p)",
     "HOTA ~72% (ByteTrack baseline)", "기존 트래커 베이스라인 평가 + HOTA/DetA/AssA 분해",
     "축구 MOT 공식 벤치마크 수립, fast motion·가림 상황 성능 부족 정량화", "직접",
     "https://arxiv.org/abs/2204.06918"),

    ("P1-2", "1. 축구 MOT", "Cui et al.", "SportsMOT: A Large Multi-Object Tracking Dataset in Multiple Sports Scenes", 2023, "ICCV 2023",
     "농구/배구/축구 포괄 대규모 MOT 데이터셋 부재, 빠른 움직임+비슷한 외관의 난제", "객체 연관(association)이 sports MOT의 핵심 난제임을 가정",
     "MixSort (MixFormer-like auxiliary association)", "SportsMOT (240 seq, 150K+ frames, 1.6M+ boxes)",
     "SportsMOT/MOT17 HOTA·IDF1 SOTA", "MixSort 제안 후 baseline들과 HOTA/IDF1 비교",
     "sports 특화 association 난제 정량 증명, 대규모 데이터셋 공개", "직접",
     "https://arxiv.org/abs/2304.05170"),

    ("P1-3", "1. 축구 MOT", "Scott, Uchida et al.", "TeamTrack: A Dataset for Multi-Sport Multi-Object Tracking in Full-pitch Videos", 2024, "CVPR 2024 Workshop (CVsports)",
     "축구/농구/핸드볼 풀 필드 전역 추적 부재", "Top-view drone + fisheye side view 결합하면 전역 추적 가능",
     "YOLOv8 + LSTM/GNN forecaster", "TeamTrack (4M+ boxes)",
     "HOTA/IDF1/MOTA 벤치마크 제공", "다중 종목·다중 시점에서 trajectory forecasting GNN 통합 평가",
     "전체 필드 MOT 벤치마크 + trajectory GNN 통합", "직접",
     "https://arxiv.org/abs/2404.13868"),

    ("P1-4", "1. 축구 MOT", "Naik et al.", "A Survey on Soccer Player Detection and Tracking with Videos", 2024, "The Visual Computer (Springer)",
     "축구 영상 탐지·추적 기법 체계적 정리 필요", "YOLO/Faster R-CNN + DeepSORT/ByteTrack 계열 정리",
     "서베이 (리뷰 논문)", "-", "-", "문헌 분류 및 기법 체계화",
     "축구 MOT 도전과제(유사 유니폼, 예측 불가 동작) 정리", "직접",
     "https://link.springer.com/article/10.1007/s00371-024-03367-6"),

    ("P1-5", "1. 축구 MOT", "Wang", "Enhancing the Performance and Accuracy in Real-Time Football and Player Detection Using Upgraded YOLOv5 Architecture", 2025, "Int. J. Computational Intelligence Systems (Springer)",
     "축구 방송 영상에서 저해상도·motion blur·occlusion으로 탐지 저하", "YOLOv5에 attention 추가 + DeepSORT 통합으로 강건성 확보",
     "YOLOv5 + Attention + DeepSORT", "축구 방송 영상",
     "선수 mAP 97.7%, 공 mAP 65.3%", "YOLOv5 + attention module 추가 후 DeepSORT 통합 평가",
     "Dual-model framework로 선수/공 동시 탐지 개선", "직접",
     "https://link.springer.com/article/10.1007/s44196-024-00565-x"),

    ("P1-6", "1. 축구 MOT", "-", "Enhanced Dual-Model Framework for Precision Player Tracking and Ball Detection in Soccer Videos", 2025, "The Visual Computer (Springer)",
     "선수와 공은 크기/속도가 달라 단일 모델 동시 탐지 시 성능 저하", "선수/공 별도 모델 + 공유 특징 추출이 동시 성능 개선",
     "Dual-model (선수용/공용 분리)", "축구 영상",
     "선수 mAP 97.7%, 공 mAP 65.3%", "별도 모델 학습 후 공유 특징 추출로 통합",
     "Dual-model 프레임워크로 정밀도 개선", "직접",
     "https://link.springer.com/article/10.1007/s00371-025-04118-x"),

    ("P1-7", "1. 축구 MOT", "-", "Self-Supervised Small Soccer Player Detection and Tracking", 2020, "ACM MMSports 2020",
     "작은 선수 객체(먼 거리) 탐지 시 라벨링 비용 큼", "자기지도 학습으로 레이블 없이 작은 선수 탐지 가능",
     "Self-supervised detector", "축구 영상", "-", "자기지도 pretext task 기반 detector 학습",
     "레이블 없는 작은 선수 탐지 가능성 제시", "간접",
     "https://dl.acm.org/doi/10.1145/3422844.3423054"),

    ("P1-8", "1. 축구 MOT", "Zhang, Wang, Wang et al.", "FairMOT: On the Fairness of Detection and Re-Identification in Multi-Object Tracking", 2021, "IJCV 2021",
     "JDE 프레임워크에서 탐지와 Re-ID 경쟁으로 학습 불공정", "Anchor-free + multi-task 공정 학습으로 양자 성능 동시 개선",
     "CenterNet + multi-layer fusion + low-dim embedding", "MOT17/20", "MOT17 MOTA 73.7, IDF1 72.3",
     "JDE vs Fair 비교 실험으로 anchor ambiguity 효과 증명", "JDE 프레임워크 공정성 확보, scale competition 완화", "간접",
     "https://github.com/ifzhang/FairMOT"),

    ("P1-9", "1. 축구 MOT", "Meinhardt, Kirillov et al.", "TrackFormer: Multi-Object Tracking with Transformers", 2022, "CVPR 2022",
     "전통적 tracking-by-detection의 motion/appearance 수동 모델링 한계", "Transformer set prediction으로 end-to-end tracking 가능",
     "Encoder-decoder Transformer + track query", "MOT17", "MOT17 MOTA 65.0%",
     "frame-to-frame set prediction, autoregressive track query 전달", "최초 end-to-end Transformer MOT", "간접",
     "https://arxiv.org/abs/2101.02702"),

    ("P1-10", "1. 축구 MOT", "Zeng et al.", "MOTR: End-to-End Multiple-Object Tracking with Transformer", 2022, "ECCV 2022",
     "완전 end-to-end MOT framework 부재", "DETR 기반 iterative track query 업데이트 + Collective Average Loss",
     "DETR + Temporal Aggregation Network", "DanceTrack, MOT17",
     "DanceTrack HOTA +6.5% vs ByteTrack; MOT17 MOTA 71.9%", "CAL + TAN 통한 temporal consistency 학습",
     "End-to-end Transformer MOT 성능 향상", "간접",
     "https://www.ecva.net/papers/eccv_2022/papers_ECCV/papers/136870648.pdf"),

    ("P1-11", "1. 축구 MOT", "Huang et al.", "Iterative Scale-Up ExpansionIoU and Deep Features Association for Multi-Object Tracking in Sports", 2024, "WACV 2024 Workshop (RWS)",
     "Sports MOT에서 IoU 기반 연관이 빠른 비선형 움직임에 취약", "Expansion IoU + deep Re-ID로 motion-agnostic 연관",
     "Deep-EIoU", "SportsMOT, SoccerNet-Tracking", "SportsMOT HOTA 77.2%, SoccerNet 85.4%",
     "Iterative scale-up EIoU 반복 매칭 + deep feature 결합", "Sports 전용 SOTA 달성, motion-agnostic 연관 기법 정립", "직접",
     "https://openaccess.thecvf.com/content/WACV2024W/RWS/papers/Huang_Iterative_Scale-Up_ExpansionIoU_and_Deep_Features_Association_for_Multi-Object_Tracking_WACVW_2024_paper.pdf"),

    ("P1-12", "1. 축구 MOT", "-", "GTATrack: Winner Solution to SoccerTrack 2025 with Deep-EIoU and Global Tracklet Association", 2025, "arXiv 2025 (SoccerTrack 2025 Winner)",
     "Fisheye 4096×1080 22명 동시·심한 distortion·빈번 가림", "Online Deep-EIoU + Offline GTA-Link tracklet clustering",
     "Deep-EIoU + GTA-Link (appearance+spatio-temporal)", "SoccerTrack, SoccerNet",
     "HOTA SORT 대비 +6.84%, Deep-EIoU 대비 +3.7%", "Online tracking 후 global offline clustering 후처리",
     "SoccerTrack 2025 우승, global tracklet association 효과 증명", "직접",
     "https://arxiv.org/html/2602.00484"),

    # ---------- Stage 2: 가림(Occlusion) 처리 — 범용 + 스포츠 ----------
    ("P2-1", "2. 가림 처리", "Bewley, Ge, Ott, Ramos, Upcroft", "Simple Online and Realtime Tracking", 2016, "ICIP 2016",
     "실시간·간결·고성능 tracking 프레임워크 필요", "Kalman filter + Hungarian algorithm만으로도 SOTA 가능",
     "Kalman Filter + Hungarian", "MOT15", "260 Hz, 당시 SOTA 대비 20배 빠름",
     "KF로 상태 예측 + IoU 기반 헝가리안 association", "MOT 분야의 고전 baseline 확립", "간접 (짧은 가림만 회복)",
     "https://arxiv.org/abs/1602.00763"),

    ("P2-2", "2. 가림 처리", "Wojke, Bewley, Paulus", "Simple Online and Realtime Tracking with a Deep Association Metric", 2017, "ICIP 2017",
     "SORT가 가림/ID switch에 취약", "CNN appearance feature로 cascade matching하면 가림 회복 가능",
     "SORT + CNN appearance", "MOT16", "ID switch 45% 감소",
     "Mahalanobis distance + cosine similarity cascade matching", "DeepSORT — 실용적 MOT 표준", "직접",
     "https://arxiv.org/abs/1703.07402"),

    ("P2-3", "2. 가림 처리", "Zhang et al.", "ByteTrack: Multi-Object Tracking by Associating Every Detection Box", 2022, "ECCV 2022",
     "가림된 객체는 detection score 낮아 고 threshold 방식에서 누락", "모든 detection box를 활용하면 가림 객체 보존 가능",
     "BYTE association", "MOT17/20, HiEve", "MOT17 MOTA 80.3, IDF1 77.3, HOTA 63.1",
     "High-score primary matching → low-score secondary matching (tracklet 재활성화)",
     "9개 SOTA tracker에 적용 시 IDF1 +1~10, 가림 회복 패러다임 확립", "직접",
     "https://arxiv.org/abs/2110.06864"),

    ("P2-4", "2. 가림 처리", "Aharon, Orfaig, Bobrovsky", "BoT-SORT: Robust Associations Multi-Pedestrian Tracking", 2022, "arXiv 2022",
     "카메라 모션이 association 성능 저하 유발", "CMC + 개선 Kalman state + Re-ID 융합으로 강건성 향상",
     "CMC + Kalman + IoU + Re-ID", "MOT17/20", "MOT17 MOTA 80.5, IDF1 80.2, HOTA 65.0",
     "카메라 모션 보상 + Kalman state 확장 + appearance 융합", "CMC 통합 SORT 계열 SOTA", "직접",
     "https://arxiv.org/abs/2206.14651"),

    ("P2-5", "2. 가림 처리", "Cao et al.", "Observation-Centric SORT: Rethinking SORT for Robust Multi-Object Tracking", 2023, "CVPR 2023",
     "가림 구간에서 Kalman 예측 오차 누적, 복구 실패", "Observation-centric re-update로 가림 이후 관측 기반 궤적 재구성",
     "OC-SORT", "MOT17/20, KITTI, DanceTrack", "다수 벤치마크 SOTA, 700+ FPS (CPU)",
     "Observation-centric re-update + virtual trajectory 재구성", "가림 구간 motion recovery 패러다임 정립", "직접 (핵심)",
     "https://arxiv.org/abs/2203.14360"),

    ("P2-6", "2. 가림 처리", "Du, Zhao, Song et al.", "StrongSORT: Make DeepSORT Great Again", 2023, "IEEE TMM 2023",
     "DeepSORT 구식화, 최신 기법 미활용", "AFLink + GSI 후처리로 DeepSORT를 SOTA로 복원",
     "DeepSORT + AFLink + GSI", "MOT17/20", "HOTA·IDF1 1위, 2위 대비 +1.3~2.2",
     "DeepSORT 백본 + AFLink offline tracklet linking + Gaussian-Smoothed Interpolation",
     "DeepSORT 계열 SOTA 복귀, offline post-processing 정립", "직접",
     "https://arxiv.org/abs/2202.13514"),

    ("P2-7", "2. 가림 처리", "Maggiolino et al.", "Deep OC-SORT: Multi-Pedestrian Tracking by Adaptive Re-identification", 2023, "arXiv 2023",
     "OC-SORT가 appearance를 활용하지 않아 장기 가림 취약", "Adaptive Re-ID weight로 appearance 동적 통합",
     "OC-SORT + adaptive Re-ID", "MOT17/20", "MOT17/20 HOTA SOTA",
     "Appearance confidence 기반 Re-ID weight 동적 조정", "장기 가림 복구 성능 개선", "직접",
     "https://arxiv.org/pdf/2302.11813"),

    ("P2-8", "2. 가림 처리", "Wang, Xiao et al.", "Repulsion Loss: Detecting Pedestrians in a Crowd", 2018, "CVPR 2018",
     "군중 속 가림된 보행자 탐지 실패", "Attraction+Repulsion 결합 손실이 인접 객체 혼동 줄임",
     "Repulsion Loss (RepGT + RepBox)", "CityPersons, CrowdHuman", "가림 상황 SOTA",
     "Attraction term + RepGT (다른 GT 멀리) + RepBox (다른 예측 멀리) 손실 결합",
     "가림된 보행자 탐지 loss-level 해법 정립", "직접",
     "https://arxiv.org/abs/1711.07752"),

    ("P2-9", "2. 가림 처리", "Chu et al.", "Occlusion Handling in Generic Object Detection: A Review", 2022, "arXiv Survey",
     "가림 처리 기법 분산, 체계적 정리 필요", "Loss/feature/part/GAN 기반 분류",
     "서베이", "-", "-", "기법 분류 (intra-class vs inter-class)",
     "가림 처리 기법 체계화", "직접",
     "https://arxiv.org/pdf/2101.08845"),

    ("P2-10", "2. 가림 처리", "-", "Occlusion Handling and Multi-Scale Pedestrian Detection Based on Deep Learning: A Review", 2022, "IEEE Access 2022",
     "보행자 탐지에서 가림/다중 스케일 문제 분산", "기법 리뷰로 체계화", "서베이", "-", "-",
     "보행자 탐지 가림 기법 정리", "보행자 탐지 서베이", "직접",
     "https://ieeexplore.ieee.org/document/9718221/"),

    ("P2-11", "2. 가림 처리", "-", "OATrack: Towards Occlusion-Aware Multi-Pedestrian Tracking", 2024, "MDPI Applied Sciences",
     "Kalman gain이 가림 상황에서 부적절하게 업데이트", "가림율 추정 후 adaptive gain/cue weight 조정",
     "Occlusion Perception Module + Adaptive Kalman", "MOT17/20", "-",
     "Occlusion rate 추정 → Kalman innovation gain suppression + association cue weight 조정",
     "Occlusion-aware Kalman + adaptive association 정립", "직접 (핵심 설계)",
     "https://www.mdpi.com/2076-3417/15/24/13045"),

    ("P2-12", "2. 가림 처리", "Zhu et al.", "Handling Heavy Occlusion in Dense Crowd Tracking by Focusing on the Heads", 2023, "arXiv 2023 / Springer 2024",
     "군중 속 몸통은 심각한 가림, 머리는 상대적으로 덜 가려짐", "Joint head-body detector로 머리 우선 추적",
     "Anchor-free joint head-body + SoftNMS", "MOT20, HT21", "MOT20·HT21 성능 향상",
     "Head + body joint detection + SoftNMS 후처리", "Dense crowd tracking head-focus 패러다임", "직접",
     "https://arxiv.org/abs/2304.07705"),

    ("P2-13", "2. 가림 처리", "-", "Occlusion-Aware Spatial Attention Transformer for Occluded Object Recognition", 2022, "Pattern Recognition Letters 2022",
     "가림 영역이 attention에 악영향", "Spatial attention + occlusion-aware loss로 가림 영역 식별·무시",
     "Transformer + Occlusion-aware Loss", "occluded object benchmark", "-",
     "Spatial attention으로 가림 영역 식별 + loss 설계", "Attention 레벨 가림 처리", "직접",
     "https://www.sciencedirect.com/science/article/abs/pii/S0167865522001581"),

    ("P2-14", "2. 가림 처리", "-", "Graph Convolution Neural Network-Based Data Association for Online Multi-Object Tracking", 2021, "IEEE Access 2021",
     "가림 시 appearance matching 불안정", "GCN으로 객체 간 affinity 추정, pose feature 활용",
     "GCN + pose feature", "MOT16/17", "MOT16 MOTA 80.6, MOT17 81.1, HOTA 65.3/65.1",
     "GCN affinity graph + pose 결합 association", "GNN 기반 가림 강건 association 확립", "직접",
     "https://ieeexplore.ieee.org/document/9514568/"),

    ("P2-15", "2. 가림 처리", "-", "Occlusion-Related Graph Convolutional Neural Network for Multi-Object Tracking", 2024, "Image and Vision Computing (Elsevier)",
     "가림 관계를 graph에 명시 안 함", "Occlusion edge attribute 도입으로 association 모델링",
     "GCN + occlusion edge", "MOT 벤치마크", "-",
     "그래프에 명시적 occlusion edge attribute 부여", "Occlusion-aware GNN association 심화", "직접",
     "https://www.sciencedirect.com/science/article/abs/pii/S0262885624004220"),

    ("P2-16", "2. 가림 처리", "Back et al.", "Unseen Object Amodal Instance Segmentation via Hierarchical Occlusion Modeling", 2022, "ICRA 2022",
     "가려진 객체의 완전 형태 복원 어려움", "Visible+Amodal+Occlusion mask 계층적 예측",
     "UOAIS (Hierarchical Occlusion Modeling)", "OCID, OSD, OSD-Amodal",
     "Amodal SOTA (로봇 tabletop 데이터)", "Visible/Amodal/Occlusion 3개 mask 계층적 학습",
     "Amodal instance segmentation 계층적 프레임워크", "직접 (가려진 부분 복원)",
     "https://arxiv.org/abs/2109.11103"),

    ("P2-17", "2. 가림 처리", "-", "SAMEO: Segment Anything, Even Occluded", 2025, "arXiv 2025",
     "Amodal segmentation 데이터 부족 + 일반화 어려움", "SAM 기반 foundation model 적응 + 300K Amodal-LVIS",
     "SAM 기반 Amodal", "Amodal-LVIS 300K", "Amodal SOTA",
     "SAM adapter + Amodal-LVIS 데이터 구축", "Foundation model 기반 Amodal 확장", "직접",
     "https://arxiv.org/html/2503.06261v1"),

    ("P2-18", "2. 가림 처리", "-", "Sequential Amodal Segmentation via Cumulative Occlusion Learning", 2024, "arXiv 2024",
     "Invisible 영역 모델링의 불확실성 포착 어려움", "Diffusion iterative refinement + cumulative mask",
     "Diffusion Model + Cumulative Mask", "Amodal dataset", "-",
     "Diffusion 모델로 iterative amodal refinement", "Diffusion 기반 amodal + uncertainty 포착", "직접",
     "https://arxiv.org/html/2405.05791v1"),

    ("P2-19", "2. 가림 처리", "Dendorfer et al.", "MOT20: A Benchmark for Multi Object Tracking in Crowded Scenes", 2020, "arXiv 2020",
     "군중 밀집 환경 MOT 벤치마크 부재", "프레임당 최대 246명 고밀도 데이터셋 제공",
     "-", "MOT20 (8 seq, 최대 246명/frame)", "-",
     "고밀도 시퀀스 수집 및 어노테이션", "군중 MOT 표준 벤치마크", "직접",
     "https://arxiv.org/abs/2003.09003"),

    ("P2-20", "2. 가림 처리", "-", "CrowdTrack: A Benchmark for Difficult Multiple Pedestrian Tracking", 2025, "arXiv 2025",
     "기존 MOT 벤치마크가 쉬운 시나리오에 편중", "난이도 높은 군중 MOT 시퀀스 제공",
     "-", "CrowdTrack", "-",
     "어려운 군중 시나리오 큐레이션", "난이도 높은 군중 벤치마크", "직접",
     "https://arxiv.org/pdf/2507.02479"),

    ("P2-21", "2. 가림 처리", "-", "MCTrack: A Unified 3D Multi-Object Tracking Framework for Autonomous Driving", 2024, "IROS 2025",
     "자율주행 3D MOT에서 BEV/image 매칭 분산", "BEV-plane 1차 + image-plane 2차 dual matching",
     "MCTrack (BEV+image dual matching)", "KITTI, nuScenes, Waymo", "KITTI/nuScenes/Waymo SOTA",
     "BEV 매칭 후 unmatched 객체를 image plane에서 보완 매칭", "자율주행 3D MOT 통합 프레임워크", "직접",
     "https://arxiv.org/abs/2409.16149"),

    # ---------- Stage 3: 등번호 인식 ----------
    ("P3-1", "3. 등번호 인식", "Liu & Bhanu", "Pose-Guided R-CNN for Jersey Number Recognition in Sports", 2019, "CVPR 2019 Workshop (CVSports)",
     "선수 자세/시점 변화로 등번호 가림·왜곡", "Pose keypoint로 ROI 정렬 후 인식하면 강건",
     "Pose-Guided R-CNN", "스포츠 영상", "-",
     "Pose keypoint 기반 등번호 ROI 정렬 후 R-CNN", "Pose-guided jersey recognition 정립", "직접",
     "https://openaccess.thecvf.com/content_CVPRW_2019/papers/CVSports/Liu_Pose-Guided_R-CNN_for_Jersey_Number_Recognition_in_Sports_CVPRW_2019_paper.pdf"),

    ("P3-2", "3. 등번호 인식", "Gerke et al.", "Soccer Jersey Number Recognition Using Convolutional Neural Networks", 2015, "ICCV Workshops 2015",
     "축구 등번호 인식 자동화 필요", "CNN 분류로 등번호 인식 가능",
     "CNN 분류기", "축구 영상", "-",
     "Cropped jersey image → CNN 분류", "초기 CNN 기반 jersey recognition", "간접",
     "https://ieeexplore.ieee.org/document/7406449/"),

    ("P3-3", "3. 등번호 인식", "Vats et al.", "Multi-task Learning for Jersey Number Recognition in Ice Hockey", 2021, "ACM MMSports 2021",
     "Jersey number 단독 학습보다 multi-task가 효과적", "번호+팀+포지션 공동 학습이 상호 보완",
     "Multi-task CNN", "Ice hockey 영상", "-",
     "번호/팀/포지션 multi-head 학습", "Multi-task 활용 jersey recognition 개선", "간접",
     "https://dl.acm.org/doi/10.1145/3475722.3482794"),

    ("P3-4", "3. 등번호 인식", "Balaji et al.", "Jersey Number Recognition using Keyframe Identification from Low-Resolution Broadcast Videos", 2023, "ACM MMSports 2023",
     "방송 영상 등번호는 극소수 프레임에서만 보임", "Keyframe identification으로 가시 프레임 선별 후 STR",
     "Keyframe ID + Scene Text Recognition", "Broadcast 영상", "정확도 +37.81%, +37.70% (두 테스트셋)",
     "Keyframe 선별 module → STR 인식", "Keyframe selection 패러다임 정립", "직접",
     "https://dl.acm.org/doi/10.1145/3606038.3616162"),

    ("P3-5", "3. 등번호 인식", "Koshkina & Elder", "A General Framework for Jersey Number Recognition in Sports Video", 2024, "arXiv 2405.13896",
     "Jersey recognition 단계들 분산", "YOLOv4 detection → localization → 4-stage STR 통합",
     "YOLOv4 + Localization + 4-stage STR", "스포츠 영상", "-",
     "Detection → localization → 4-stage text recognition pipeline", "범용 jersey recognition 프레임워크", "간접",
     "https://arxiv.org/abs/2405.13896"),

    ("P3-6", "3. 등번호 인식", "Grad et al.", "Single-Stage Uncertainty-Aware Jersey Number Recognition in Soccer", 2025, "CVPR 2025 Workshop (CVSports)",
     "2-stage detect-recognize 비효율 + 불확실성 미반영", "단일 단계 통합 + uncertainty estimation",
     "Single-Stage + Uncertainty", "SoccerNet Jersey", "-",
     "Detection+recognition 단일 단계 + uncertainty 추정", "Uncertainty-aware jersey recognition 도입", "직접",
     "https://openaccess.thecvf.com/content/CVPR2025W/CVSPORTS/papers/Grad_Single-Stage_Uncertainty-Aware_Jersey_Number_Recognition_in_Soccer_CVPRW_2025_paper.pdf"),

    ("P3-7", "3. 등번호 인식", "SoccerNet 2023 participants", "SoccerNet 2023 Challenges Results (Jersey Number Recognition track)", 2023, "arXiv 2309.06006",
     "Jersey recognition 벤치마크 및 상위 솔루션 비교 필요", "Tracklet-level majority voting + STR이 최적",
     "Tracklet majority voting + STR", "SoccerNet Jersey",
     "Test 90.09%, Challenge 90.95%", "각 track 상위 솔루션 결과 보고",
     "SoccerNet Jersey 2023 SOTA 정리", "직접",
     "https://arxiv.org/pdf/2309.06006"),

    ("P3-8", "3. 등번호 인식", "Vats et al.", "Ice Hockey Player Identification via Transformers and Weakly Supervised Learning", 2022, "CVPR 2022 Workshop",
     "Tracklet-level 선수 ID에 Transformer 활용 미진", "Weakly supervised Transformer가 tracklet에서 효과적",
     "Weakly-supervised Transformer", "Ice hockey tracklet", "-",
     "Tracklet을 Transformer에 입력, weak label로 학습",
     "Transformer 기반 tracklet ID 가능성 제시", "간접",
     "-"),

    ("P3-9", "3. 등번호 인식", "-", "Spatio-Temporal Jersey Number Recognition (ResNet+LSTM)", 2021, "SoccerNet baseline",
     "Single-frame 인식 한계", "Tracklet temporal aggregation으로 강건성 확보",
     "ResNet + LSTM", "SoccerNet Jersey", "-",
     "ResNet backbone + LSTM temporal aggregation", "Temporal aggregation baseline", "간접",
     "-"),

    # ---------- Stage 4: Re-ID + 통합 파이프라인 ----------
    ("P4-1", "4. Re-ID + 통합", "Senocak → Li et al.", "Multi-camera Multi-player Tracking with Deep Player Identification in Sports Video", 2020, "Pattern Recognition 2020",
     "다중 카메라 선수 ID 유지 어려움", "딥 Re-ID embedding으로 다중 카메라 ID 일관성 확보",
     "Deep Re-ID embedding", "Multi-camera sports",
     "-", "다중 카메라 영상에서 embedding matching",
     "Multi-camera player ID 초기 연구", "간접",
     "https://dl.acm.org/doi/10.1016/j.patcog.2020.107260"),

    ("P4-2", "4. Re-ID + 통합", "Habel et al.", "Sports Re-ID: Improving Re-Identification Of Players In Broadcast Videos Of Team Sports", 2022, "arXiv 2206.02373",
     "방송 team sport에서 Re-ID 어려움", "Part-based embedding + team-aware sampling",
     "Part-based + team-aware", "Sports broadcast", "-",
     "Part-based embedding + team-aware 샘플링 학습",
     "Team sport Re-ID 개선", "직접",
     "https://arxiv.org/abs/2206.02373"),

    ("P4-3", "4. Re-ID + 통합", "SoccerNet Re-ID 2022 Winner", "Multi-task Learning for Joint Re-identification, Team Affiliation, and Role Classification for Sports Visual Tracking", 2023, "ACM MMSports 2023 / arXiv 2401.09942",
     "Re-ID 단독 학습 한계", "Jersey+team+pose 공동 학습이 Re-ID 성능 향상",
     "Multi-task (Re-ID + team + role + pose)", "SoccerNet Re-ID",
     "mAP 86.0, Rank-1 81.5", "Multi-task head 공동 학습",
     "SoccerNet Re-ID 2022 우승, multi-task 효과", "직접",
     "https://arxiv.org/html/2401.09942v1"),

    ("P4-4", "4. Re-ID + 통합", "Akan & Varlı", "Reidentifying Soccer Players in Broadcast Videos Using Body Feature Alignment Based on Pose", 2023, "CNIOT 2023",
     "Soccer broadcast Re-ID에서 pose 변화 강건성 부족", "Pose 기반 body part alignment로 part-wise 임베딩",
     "Pose + Part alignment", "Soccer broadcast", "-",
     "Pose estimation → body part alignment → part-wise embedding",
     "Pose-based soccer Re-ID 개선", "직접",
     "https://dl.acm.org/doi/abs/10.1145/3603781.3603860"),

    ("P4-5", "4. Re-ID + 통합", "Miao et al.", "Pose-Guided Feature Alignment for Occluded Person Re-Identification", 2019, "ICCV 2019",
     "가림 시 person Re-ID 성능 저하", "Pose keypoint로 가림 영역 attention 억제",
     "PGFA", "Occluded-Duke, Partial-REID", "가림 Re-ID SOTA (당시)",
     "Pose keypoint 기반 가림 영역 마스킹 + 나머지 part attention",
     "Occluded person Re-ID 기초 정립", "직접",
     "https://openaccess.thecvf.com/content_ICCV_2019/html/Miao_Pose-Guided_Feature_Alignment_for_Occluded_Person_Re-Identification_ICCV_2019_paper.html"),

    ("P4-6", "4. Re-ID + 통합", "Somers et al.", "SoccerNet Game State Reconstruction: End-to-End Athlete Tracking and Identification on a Minimap", 2024, "CVPR 2024 Workshop (CVsports)",
     "단일 방송 카메라 → 2D minimap 전체 재구성 부재", "탐지+추적+번호+팀+역할 통합하면 GSR 가능",
     "YOLO + tracker + Re-ID + jersey + team", "SoccerNet-GSR", "GS-HOTA 제공",
     "End-to-end pipeline 구성 및 minimap 재구성", "GSR 벤치마크 도입", "직접",
     "https://openaccess.thecvf.com/content/CVPR2024W/CVsports/papers/Somers_SoccerNet_Game_State_Reconstruction_End-to-End_Athlete_Tracking_and_Identification_on_CVPRW_2024_paper.pdf"),

    ("P4-7", "4. Re-ID + 통합", "SoccerNet 2025 GSR Winner", "From Broadcast to Minimap: Achieving State-of-the-Art SoccerNet Game State Reconstruction", 2025, "CVPR 2025",
     "GSR 파이프라인 각 모듈 통합 최적화 필요", "모든 최신 기법 통합이 SOTA 달성",
     "YOLOv5m + SegFormer + DeepSORT + Re-ID + Jersey", "SoccerNet-GSR", "GS-HOTA SOTA",
     "전 모듈 최신 기법 통합 및 파인튜닝", "SoccerNet 2025 GSR SOTA", "직접",
     "https://cvpr.thecvf.com/virtual/2025/35501"),

    ("P4-8", "4. Re-ID + 통합", "-", "Occluded Person Re-Identification with Deep Learning: A Survey and Perspectives", 2023, "arXiv 2311.00603",
     "Occluded Re-ID 기법 분산", "Part/pose/semantic/Transformer 기반 체계화",
     "서베이", "-", "-",
     "기법별 분류 및 성능 비교",
     "Occluded Re-ID 서베이", "직접",
     "https://arxiv.org/pdf/2311.00603"),

    ("P4-9", "4. Re-ID + 통합", "Koshkina et al.", "Contrastive Learning for Sports Video: Unsupervised Player Classification", 2021, "CVPR 2021 Workshop (CVSports)",
     "팀 분류 레이블 비용 큼", "Contrastive learning으로 팀 간/팀 내 분리 가능",
     "Contrastive (SimCLR-like)", "Soccer broadcast",
     "단일 프레임 94%, 500 프레임 97% 팀 분류",
     "Contrastive loss로 팀 embedding 학습",
     "Unsupervised 팀 분류 가능성 증명", "간접",
     "https://arxiv.org/abs/2104.10068"),
]

# ---------- 엑셀 생성 ----------
wb = Workbook()
ws = wb.active
ws.title = "Literature Review"

headers = [
    "논문ID", "Stage", "저자", "논문 제목", "출판연도", "학회/저널",
    "문제 정의", "가설/접근법", "모델/아키텍처", "데이터셋",
    "핵심 성능 지표", "실험 방법 요약", "결론/기여점",
    "가림(Occlusion) 관련성", "DOI/URL"
]

# 헤더
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 데이터
stage_colors = {
    "1. 축구 MOT": "DDEBF7",
    "2. 가림 처리": "FFF2CC",
    "3. 등번호 인식": "E2EFDA",
    "4. Re-ID + 통합": "FCE4D6",
}
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, paper in enumerate(papers, start=2):
    stage = paper[1]
    fill_color = stage_colors.get(stage, "FFFFFF")
    for col_idx, value in enumerate(paper, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.font = Font(size=10)

# 열 너비
col_widths = {
    "A": 9,   # 논문ID
    "B": 14,  # Stage
    "C": 22,  # 저자
    "D": 50,  # 제목
    "E": 10,  # 연도
    "F": 26,  # 학회/저널
    "G": 38,  # 문제정의
    "H": 38,  # 가설
    "I": 30,  # 모델
    "J": 28,  # 데이터셋
    "K": 30,  # 성능
    "L": 38,  # 실험방법
    "M": 38,  # 결론
    "N": 14,  # 가림관련성
    "O": 55,  # URL
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# 행 높이
ws.row_dimensions[1].height = 36
for row_idx in range(2, len(papers) + 2):
    ws.row_dimensions[row_idx].height = 120

# 헤더 행 고정
ws.freeze_panes = "A2"
# 필터
ws.auto_filter.ref = f"A1:O{len(papers)+1}"

# 요약 시트
ws2 = wb.create_sheet("Summary")
ws2.append(["Phase 1 문헌 탐색 결과 요약"])
ws2.append([])
ws2.append(["총 논문 수", len(papers)])
ws2.append([])
ws2.append(["Stage별 분포"])
from collections import Counter
cnt = Counter([p[1] for p in papers])
for stage, n in sorted(cnt.items()):
    ws2.append([stage, n])

ws2.append([])
ws2.append(["연도별 분포"])
year_cnt = Counter([p[4] for p in papers])
for year in sorted(year_cnt.keys()):
    ws2.append([year, year_cnt[year]])

ws2.append([])
ws2.append(["가림 관련성"])
occ_cnt = Counter([p[13] for p in papers])
for k, v in occ_cnt.items():
    ws2.append([k, v])

ws2["A1"].font = Font(bold=True, size=14)
for row in [3, 5, 12, 21]:
    for col in ["A"]:
        c = ws2[f"{col}{row}"] if row <= ws2.max_row else None
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 12

# 저장
output_path = r"C:\Users\neohc\Desktop\ClaudeCode\YoungScientist\_workspace\02_literature_table.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Total papers: {len(papers)}")
print(f"Stage distribution: {dict(cnt)}")
