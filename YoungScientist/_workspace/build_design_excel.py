"""
Phase 3 연구 설계서의 핵심 데이터를 엑셀로 정리
출력: 06_research_design_table.xlsx (6 시트)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols):
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def fill_data(ws, data, start_row=2, colors=None):
    for row_idx, row in enumerate(data, start=start_row):
        fill = colors[(row_idx - start_row) % len(colors)] if colors else "FFFFFF"
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            c.font = Font(size=10)

# =========================
# Sheet 1: Research Questions & Hypotheses
# =========================
ws1 = wb.active
ws1.title = "1. RQ & Hypotheses"

rq_headers = ["ID", "유형", "진술 / 가설", "검증 지표", "유의 수준 / 검정", "관련 공백"]
for col_idx, h in enumerate(rq_headers, start=1):
    ws1.cell(row=1, column=col_idx, value=h)
style_header(ws1, len(rq_headers))

rq_data = [
    ("PRQ", "Primary RQ", "단일 카메라 축구 영상에서 amodal + uniform-aware Re-ID + uncertainty propagation 통합 프레임워크가 가림(≥50%) 구간 선수 식별 성능을 SOTA 대비 IDF1 ≥5%, Jersey ≥3% 개선할 수 있는가?", "IDF1, Jersey Accuracy, Occluded-HOTA", "paired t-test, p<0.05", "①②③"),
    ("SRQ1", "Secondary RQ", "SAMEO pseudo-label + UOAIS 계층적 학습으로 축구 선수의 amodal mask를 얼마나 정확히 예측할 수 있는가?", "amodal mIoU, visible-amodal IoU gap", "-", "①"),
    ("SRQ2", "Secondary RQ", "Jersey-region branch + intra-team contrastive loss Uniform-aware Re-ID가 기존 Sports Re-ID 대비 intra-team Rank-1을 얼마나 개선하는가?", "Intra-team Rank-1, mAP", "paired t-test", "②"),
    ("SRQ3", "Secondary RQ", "Amodal→Re-ID→Jersey uncertainty graph propagation이 ECE와 AURC를 유의하게 낮추는가?", "ECE, AURC, Brier", "McNemar's test", "③"),
    ("SRQ4", "Secondary RQ (보조)", "Uncertainty-weighted GTA-Link가 45분 long-clip에서 ID consistency를 얼마나 유지하는가?", "long-IDF1, ID Switch count", "Wilcoxon", "④"),
    ("SRQ5", "Secondary RQ", "제안 프레임워크 각 모듈(amodal, uniform, uncertainty)의 개별 기여도는?", "Ablation deltas", "-", "①②③"),
    ("H1", "Hypothesis", "Soccer-Amodal Head는 가림율 30%+ 선수에 대해 amodal mIoU ≥ 0.65 달성", "amodal mIoU, 구간별", "p<0.05", "①"),
    ("H2", "Hypothesis", "Uniform-aware Re-ID는 Sports Re-ID Multi-task 대비 intra-team Rank-1 +3%p 이상 향상", "Rank-1, mAP", "paired t-test, p<0.05", "②"),
    ("H3", "Hypothesis", "Uncertainty propagation은 ECE를 0.10 이하로 감소 (baseline > 0.15)", "ECE, AURC, Brier", "McNemar's test", "③"),
    ("H4", "Hypothesis", "전체 프레임워크는 SoccerNet-Tracking에서 Deep-EIoU+GTA 대비 HOTA +2%, IDF1 +3%, 심각 가림(50~80%) 구간 IDF1 +5~10%", "HOTA, IDF1, Occluded-HOTA", "paired t-test, 95% CI", "①②③"),
    ("H5", "Hypothesis", "Amodal 모듈 제거 시 Occluded-HOTA 3%p 이상 감소", "Occluded-HOTA delta", "ablation test", "①"),
    ("H6", "Hypothesis", "Intra-team contrastive loss는 Rank-1 intra-team +2%p 이상 향상", "Rank-1 intra-team", "paired t-test", "②"),
    ("H7", "Hypothesis", "45분 long-clip에서 ID Switches를 Deep-EIoU+GTA 대비 30% 이상 감소", "ID Switch count, long-IDF1", "Wilcoxon", "④"),
]
fill_data(ws1, rq_data, colors=["DDEBF7", "E2EFDA", "FFF2CC", "FCE4D6"])
for col, w in {"A":7, "B":18, "C":65, "D":32, "E":24, "F":12}.items():
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 32
for r in range(2, len(rq_data) + 2):
    ws1.row_dimensions[r].height = 68
ws1.freeze_panes = "A2"

# =========================
# Sheet 2: Variables
# =========================
ws2 = wb.create_sheet("2. Variables")
v_headers = ["유형", "변수 ID", "변수명", "수준/단위", "의미"]
for col_idx, h in enumerate(v_headers, start=1):
    ws2.cell(row=1, column=col_idx, value=h)
style_header(ws2, len(v_headers))

v_data = [
    ("독립 IV", "IV1", "Amodal module", "{off, UOAIS-head, SAMEO-adapt, Diffusion-offline}", "Amodal 방식"),
    ("독립 IV", "IV2", "Uniform-aware Re-ID", "{baseline, +jersey region, +intra-team CL, full}", "Re-ID 확장 수준"),
    ("독립 IV", "IV3", "Uncertainty propagation", "{off, amodal→Re-ID, amodal→Re-ID→Jersey, full graph}", "전파 범위"),
    ("독립 IV", "IV4", "Occlusion rate", "{0-20%, 20-50%, 50-80%, 80%+}", "가림율 구간"),
    ("독립 IV", "IV5", "Intra-team contrastive", "{off, on}", "팀 내 contrastive 유무"),
    ("독립 IV", "IV6", "Amodal pseudo-label 검수 비율", "{0%, 10%, 30%, 100%}", "수동 검수 비율"),
    ("종속 DV", "DV1", "HOTA", "[0,1]", "TrackEval 공식"),
    ("종속 DV", "DV2", "IDF1", "[0,1]", "TrackEval"),
    ("종속 DV", "DV3", "ID Switches", "count", "TrackEval"),
    ("종속 DV", "DV4", "AssA / DetA", "[0,1]", "HOTA 분해"),
    ("종속 DV", "DV5", "MOTA", "[-∞,1]", "TrackEval"),
    ("종속 DV", "DV6", "Jersey Accuracy (top-1/3)", "[0,1]", "Tracklet level"),
    ("종속 DV", "DV7", "Re-ID mAP, Rank-1", "[0,1]", "표준 Re-ID"),
    ("종속 DV", "DV8", "Occluded-HOTA (신규)", "[0,1]", "가림 50%+ 구간"),
    ("종속 DV", "DV9", "Recovery rate (신규)", "[0,1]", "가림 복귀 후 ID 복원율"),
    ("종속 DV", "DV10", "ECE, AURC, Brier (신규)", "[0,1]", "Calibration"),
    ("종속 DV", "DV11", "amodal mIoU", "[0,1]", "COCO-style"),
    ("통제 CV", "CV1", "Detector backbone", "YOLOv8-X", "-"),
    ("통제 CV", "CV2", "Input resolution", "1280", "SoccerNet 표준"),
    ("통제 CV", "CV3", "FPS", "25", "SoccerNet 표준"),
    ("통제 CV", "CV4", "Data split", "SoccerNet 공식", "-"),
    ("통제 CV", "CV5", "Optimizer", "AdamW, lr=1e-4, cosine", "-"),
    ("통제 CV", "CV6", "Seed", "42, 123, 2024 (3회)", "반복 실험"),
    ("통제 CV", "CV7", "Pose estimator", "HRNet-W48", "SoccerNet finetune"),
    ("통제 CV", "CV8", "Hardware", "NVIDIA A100 40GB", "동일 GPU"),
]
var_colors = {"독립 IV": "DDEBF7", "종속 DV": "E2EFDA", "통제 CV": "FFF2CC"}
for row_idx, v in enumerate(v_data, start=2):
    fill = var_colors.get(v[0], "FFFFFF")
    for col_idx, val in enumerate(v, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.font = Font(size=10)
for col, w in {"A":10, "B":7, "C":28, "D":42, "E":24}.items():
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "A2"

# =========================
# Sheet 3: Experiment Design (Baselines + Ablation + Scenarios)
# =========================
ws3 = wb.create_sheet("3. Experiment Design")
row = 1
ws3.cell(row=row, column=1, value="3.1 Baselines (비교 대상 8개)").font = Font(bold=True, size=12, color="FFFFFF")
ws3.cell(row=row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1
for h_idx, h in enumerate(["ID", "Baseline", "카테고리", "원 논문", "비고"], start=1):
    c = ws3.cell(row=row, column=h_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
row += 1
baselines = [
    ("B1", "ByteTrack", "Association", "P2-3", "기본 강력 baseline"),
    ("B2", "OC-SORT", "Motion", "P2-5", "Observation-centric"),
    ("B3", "Deep OC-SORT", "Motion + Re-ID", "P2-7", "Adaptive Re-ID"),
    ("B4", "StrongSORT + AFLink + GSI", "Post-processing", "P2-6", "Offline linking"),
    ("B5", "BoT-SORT", "Motion + CMC", "P2-4", "Camera motion"),
    ("B6", "Deep-EIoU + GTA-Link", "Sports SOTA", "P1-11 + P1-12", "SoccerTrack 2025 우승"),
    ("B7", "SoccerNet GSR 2025 Winner", "Integrated", "P4-7", "SoccerNet 2025 SOTA"),
    ("B8", "Sports Re-ID Multi-task", "Re-ID", "P4-3", "SoccerNet Re-ID 2022 우승"),
]
for b in baselines:
    for c_idx, val in enumerate(b, start=1):
        cell = ws3.cell(row=row, column=c_idx, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.font = Font(size=10)ㄹ
    row += 1

# 3.2 Ablation
row += 1
ws3.cell(row=row, column=1, value="3.2 Ablation Study (10개)").font = Font(bold=True, size=12, color="FFFFFF")
ws3.cell(row=row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1
for h_idx, h in enumerate(["ID", "Ablation 변형", "제거/변경 대상", "관련 가설", "기대"], start=1):
    c = ws3.cell(row=row, column=h_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
row += 1
ablations = [
    ("A1", "w/o Amodal module", "Soccer-Amodal Head 제거", "H5", "Occluded-HOTA -3%p"),
    ("A2", "w/o Uniform-aware Re-ID", "Jersey region + intra-team CL 제거", "H2, H6", "intra-team Rank-1 -3%p"),
    ("A3", "w/o Uncertainty propagation", "Graph propagation 제거", "H3", "ECE 상승"),
    ("A4", "w/o Intra-team CL only", "contrastive loss만 제거", "H6", "Rank-1 intra-team -2%p"),
    ("A5", "Amodal: UOAIS", "Amodal head 구조 1", "-", "SAMEO 대비 비교"),
    ("A6", "Amodal: SAMEO-adapt", "Amodal head 구조 2", "-", "-"),
    ("A7", "Amodal: Diffusion", "Amodal head 구조 3", "-", "Offline만"),
    ("A8", "Pseudo-label 검수 0%", "manual curation 0%", "-", "Noisy label 영향"),
    ("A9", "Pseudo-label 검수 30%", "manual curation 30%", "-", "Cost-performance"),
    ("A10", "w/o Uncertainty-weighted GTA", "GTA 기본 distance", "H7", "Long-IDF1 -2%p"),
]
for a in ablations:
    for c_idx, val in enumerate(a, start=1):
        cell = ws3.cell(row=row, column=c_idx, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.font = Font(size=10)
    row += 1

# 3.3 Scenarios
row += 1
ws3.cell(row=row, column=1, value="3.3 실험 시나리오 (5개)").font = Font(bold=True, size=12, color="FFFFFF")
ws3.cell(row=row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1
for h_idx, h in enumerate(["ID", "시나리오", "데이터", "주 측정 지표", "검증 가설"], start=1):
    c = ws3.cell(row=row, column=h_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
row += 1
scenarios = [
    ("S1", "Broadcast (방송 영상)", "SoccerNet-Tracking broadcast", "HOTA, IDF1, MOTA, Jersey acc", "H4"),
    ("S2", "Static/Fixed camera", "SoccerNet-Tracking static or fisheye", "HOTA, IDF1", "H4"),
    ("S3", "가림율 구간별", "SoccerNet (0-20, 20-50, 50-80, 80%+)", "Occluded-HOTA, Recovery rate", "H5"),
    ("S4", "Long-clip (45분)", "SoccerNet 1×45min full", "long-IDF1, ID Switch", "H7"),
    ("S5", "Cross-domain 일반화", "MOT17/20 (일반화 확인)", "HOTA, IDF1", "-"),
]
for s in scenarios:
    for c_idx, val in enumerate(s, start=1):
        cell = ws3.cell(row=row, column=c_idx, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        cell.font = Font(size=10)
    row += 1

for col, w in {"A":8, "B":28, "C":30, "D":30, "E":30}.items():
    ws3.column_dimensions[col].width = w

# =========================
# Sheet 4: Roadmap
# =========================
ws4 = wb.create_sheet("4. Roadmap (12M)")
r_headers = ["Phase", "기간", "주요 작업", "산출물", "IMRaD 매핑"]
for col_idx, h in enumerate(r_headers, start=1):
    ws4.cell(row=1, column=col_idx, value=h)
style_header(ws4, len(r_headers))
roadmap = [
    ("P1", "M1~M2", "기반 구축 — 코드베이스, 데이터 전처리, baseline 8개 재현", "Baseline reproduction report", "Methods(데이터 준비)"),
    ("P2", "M3~M4", "Soccer-Amodal Head 개발 — SAMEO pseudo-label → UOAIS finetune", "Amodal module v1", "Methods(모듈 1)"),
    ("P3", "M5~M6", "Uniform-aware Re-ID 개발 — part + pose + jersey region + intra-team CL", "Re-ID module v1", "Methods(모듈 2)"),
    ("P4", "M7~M8", "Uncertainty propagation 통합 + end-to-end 학습", "Full pipeline v1", "Methods(모듈 3)"),
    ("P5", "M9~M10", "전체 실험 + ablation + statistical test", "실험 결과 + 분석 figure", "Results + Discussion"),
    ("P6", "M11~M12", "IMRaD 논문 작성 + 학회 투고 + 리뷰 대응", "논문 draft → final → camera-ready", "Full paper"),
]
fill_data(ws4, roadmap, colors=["DDEBF7"])
for col, w in {"A":6, "B":10, "C":55, "D":40, "E":22}.items():
    ws4.column_dimensions[col].width = w
ws4.row_dimensions[1].height = 28
for r in range(2, len(roadmap) + 2):
    ws4.row_dimensions[r].height = 48

# =========================
# Sheet 5: Risks
# =========================
ws5 = wb.create_sheet("5. Risks")
rk_headers = ["ID", "리스크", "영향", "확률", "대응 전략"]
for col_idx, h in enumerate(rk_headers, start=1):
    ws5.cell(row=1, column=col_idx, value=h)
style_header(ws5, len(rk_headers))
risks = [
    ("R1", "SAMEO pseudo-label 품질 낮음", "★★★", "중", "UOAIS/SAMEO consensus filtering + 부분 수동 검수 + self-training 반복"),
    ("R2", "Amodal annotation 비용 과다", "★★★", "중", "SoccerNet 10~30%만 수동 검수, 나머지 pseudo + curriculum"),
    ("R3", "Uniform-aware Re-ID 학습 불안정", "★★", "중", "Curriculum learning + warm-up + batch-hard mining"),
    ("R4", "Uncertainty calibration 어려움", "★★", "중", "Temperature scaling + post-hoc isotonic + Brier 모니터링"),
    ("R5", "GPU 자원 부족 (A100×4)", "★★", "중", "모듈별 단계 학습 + checkpoint 분리 + mixed precision"),
    ("R6", "Long-clip annotation 부족", "★★", "고", "SoccerNet 1×45min clip + 자체 부분 annotation"),
    ("R7", "Detection FP/FN 전파", "★★", "저~중", "YOLOv8 Repulsion tuning + dual threshold"),
    ("R8", "Re-ID intra-team gap 미해소", "★★★", "저", "Jersey region 가중치 상향 + number-conditioned embedding"),
    ("R9", "Pose estimator 실패율 축구에서 높음", "★★", "저", "HRNet SoccerNet fine-tune + occluded-pose data augmentation"),
    ("R10", "Amodal inference 속도 병목", "★★", "중", "Offline 구성 + RoI 단위 amodal 생략(가림율 낮음)"),
]
fill_data(ws5, risks, colors=["FCE4D6"])
for col, w in {"A":5, "B":36, "C":8, "D":6, "E":55}.items():
    ws5.column_dimensions[col].width = w
ws5.row_dimensions[1].height = 28
for r in range(2, len(risks) + 2):
    ws5.row_dimensions[r].height = 42

# =========================
# Sheet 6: Venue Strategy
# =========================
ws6 = wb.create_sheet("6. 학회 전략")
v_headers = ["학회/저널", "마감 (추정)", "Track", "적합도", "우선순위", "비고"]
for col_idx, h in enumerate(v_headers, start=1):
    ws6.cell(row=1, column=col_idx, value=h)
style_header(ws6, len(v_headers))
venues = [
    ("CVPR 2027", "~Nov 2026", "CVSports Workshop", "★★★★★", "1순위", "Sports CV 전문 트랙"),
    ("ICCV 2027", "~Mar 2027", "Main / Sports", "★★★★", "2순위", "확장 버전 target"),
    ("ECCV 2026", "~Mar 2026", "Workshop", "★★★★", "빠른 경로", "스케줄 타이트"),
    ("WACV 2027", "~Aug 2026", "RWS Workshop", "★★★★", "Application", "실용성 강조"),
    ("BMVC 2026", "~Jul 2026", "Main", "★★★", "중간 투고", "-"),
    ("ACM MMSports 2026", "~Jun 2026", "Main", "★★★★", "Sports 전문", "Multimedia 관점"),
    ("AAAI 2027", "~Aug 2026", "Main", "★★★", "AI 일반", "-"),
    ("Pattern Recognition (Journal)", "상시", "Journal", "★★★★", "Journal 버전", "확장"),
    ("IEEE TMM (Journal)", "상시", "Journal", "★★★", "Journal 버전", "Multimedia"),
]
fill_data(ws6, venues, colors=["E2EFDA"])
for col, w in {"A":26, "B":14, "C":22, "D":10, "E":14, "F":28}.items():
    ws6.column_dimensions[col].width = w
ws6.row_dimensions[1].height = 28

# =========================
# Sheet 7: Summary
# =========================
ws7 = wb.create_sheet("7. Summary")
ws7["A1"] = "Phase 3 연구 설계서 요약 (AmoUni-SoccerTrack)"
ws7["A1"].font = Font(bold=True, size=14)
ws7.append([])
ws7.append(["연구명", "AmoUni-SoccerTrack"])
ws7.append(["연구 방향", "Amodal + Uniform-aware + Uncertainty Propagation (조합 3)"])
ws7.append(["핵심 공백 해결", "① 축구 amodal / ② uniform-aware Re-ID / ③ tracklet uncertainty"])
ws7.append(["핵심 모듈", "3개 (Soccer-Amodal Head / Uniform-aware Re-ID / Uncertainty Propagation)"])
ws7.append(["Primary RQ", "가림 ≥50% 구간에서 IDF1 ≥5%, Jersey ≥3% 개선 가능한가?"])
ws7.append(["가설 수", "H1~H7 (7개)"])
ws7.append(["Baseline 수", "8개 (B1~B8)"])
ws7.append(["Ablation 수", "10개 (A1~A10)"])
ws7.append(["실험 시나리오", "5개 (Broadcast/Static/가림구간/Long-clip/Cross-domain)"])
ws7.append(["로드맵", "12개월 (P1~P6)"])
ws7.append(["리스크", "10개 (R1~R10)"])
ws7.append(["1순위 학회", "CVPR 2027 CVSports Workshop"])
ws7.append(["예상 기여", "SoccerNet 심각 가림 IDF1 +5~10%, 축구 최초 amodal 벤치마크 공개"])
for col in ["A", "B"]:
    ws7.column_dimensions[col].width = 60

out = r"C:\Users\neohc\Desktop\ClaudeCode\YoungScientist\_workspace\06_research_design_table.xlsx"
wb.save(out)
print(f"Saved: {out}")
